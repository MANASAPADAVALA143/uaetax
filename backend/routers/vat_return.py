"""VAT Return Generator API Router"""
import hashlib
import hmac
import os
import secrets
import sys
from datetime import date, datetime
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Depends, Header, HTTPException, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, case
from pydantic import BaseModel, Field, ValidationError
import pandas as pd
import json
import io
from anthropic import Anthropic
from dotenv import load_dotenv
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Add parent directory to path for RAG import
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import get_db
from middleware.auth import get_current_company_id
from models import (
    Transaction,
    Company,
    VATReturn,
    ReconciliationResult,
    FTASubmissionLog,
    AuditLog,
)

load_dotenv()

router = APIRouter(prefix="/api/vat", tags=["VAT Return"])

# Initialize Claude client for recommendations
anthropic_api_key = os.getenv("ANTHROPIC_API_KEY")
claude_client = Anthropic(api_key=anthropic_api_key) if anthropic_api_key else None


# Pydantic models
class GenerateReturnRequest(BaseModel):
    period_start: date = Field(..., description="Period start date")
    period_end: date = Field(..., description="Period end date")


class VATReturnResponse(BaseModel):
    return_id: int
    company_id: int
    period_start: date
    period_end: date
    box1_standard_rated_supplies: float
    box2_vat_on_supplies: float
    box3_zero_rated_supplies: float
    box4_exempt_supplies: float
    box5_total_taxable_supplies: float
    box6_taxable_expenses: float
    box7_vat_on_expenses: float
    box8_vat_payable_or_refundable: float
    status: str
    created_at: datetime


class ReconciliationResponse(BaseModel):
    status: str
    difference_aed: float
    mismatches: List[Dict[str, Any]]
    recommendation: str


class VATSubmitResponse(BaseModel):
    return_id: int
    submission_status: str
    submitted_at: Optional[datetime] = None
    fta_reference_number: Optional[str] = None
    submission_error: Optional[str] = None


def _transaction_side(t: Transaction) -> str:
    return (getattr(t, "transaction_type", None) or "sale").lower()


_ENTERTAINMENT_KWS = {
    "restaurant", "hotel", "entertainment", "meals", "leisure", "recreation",
    "hospitality", "catering", "food", "beverage", "drinks", "cafe", "dinner",
    "lunch", "breakfast", "nobu", "four seasons", "hilton", "hyatt", "marriott",
    "gala", "buffet",
}


def _is_entertainment(t: Transaction) -> bool:
    """Art.53(1)(b): input VAT on entertainment/meals is not recoverable."""
    desc = (t.description or "").lower()
    return any(kw in desc for kw in _ENTERTAINMENT_KWS)


def calculate_vat_return_boxes(transactions: List[Transaction]) -> Dict[str, float]:
    """
    Calculate all 8 FTA VAT return boxes from transactions.

    UAE VAT Law (Federal Decree-Law No. 8/2017):
    - Art. 48: Reverse Charge Mechanism — buyer self-assesses VAT on
      imported services / goods from non-registered overseas suppliers.
      RC VAT appears on BOTH the output side (Box 2) and input side (Box 7)
      simultaneously.  The net cash effect is zero when the buyer is fully
      taxable, but both boxes must be populated for FTA audit purposes.
    - Box 1: Standard-rated taxable supplies (sales) — net amount only.
    - Box 2: Output VAT = (Box 1 × 5%) + RC self-assessed output VAT.
    - Box 6: Taxable expenses (standard-rated + reverse-charge purchases).
    - Box 7: Input VAT = (std purchases × 5%) + RC self-assessed input VAT.
      Entertainment expenses blocked under Art. 53 are excluded.
    """
    # ── Split by side and treatment ──────────────────────────────────────────
    sales_std  = [t for t in transactions if _transaction_side(t) == "sale"
                  and (t.vat_treatment or "") == "standard_rated"]
    sales_zero = [t for t in transactions if _transaction_side(t) == "sale"
                  and (t.vat_treatment or "") == "zero_rated"]
    sales_ex   = [t for t in transactions if _transaction_side(t) == "sale"
                  and (t.vat_treatment or "") == "exempt"]

    _purch_std_all = [t for t in transactions if _transaction_side(t) == "purchase"
                      and (t.vat_treatment or "") == "standard_rated"]
    purch_rc       = [t for t in transactions if _transaction_side(t) == "purchase"
                      and (t.vat_treatment or "") == "reverse_charge"]

    # Art.53(1)(b): split standard-rated purchases into claimable vs blocked
    purch_std_entertainment = [t for t in _purch_std_all if _is_entertainment(t)]
    purch_std               = [t for t in _purch_std_all if not _is_entertainment(t)]

    def _amt(lst):  return sum(t.amount_aed or 0 for t in lst)

    # ── Boxes 1-5: Sales side ────────────────────────────────────────────────
    box1 = _amt(sales_std)                     # Net standard-rated sales
    box3 = _amt(sales_zero)                    # Net zero-rated sales
    box4 = _amt(sales_ex)                      # Exempt sales
    box5 = box1 + box3 + box4                  # Total taxable supplies

    # RC self-assessed output VAT (Art. 48) — must be added to output side
    rc_net = _amt(purch_rc)
    rc_vat = rc_net * 0.05

    # Box 2: Output VAT = standard-rated sales VAT + RC self-assessed output
    box2 = box1 * 0.05 + rc_vat

    # ── Boxes 6-7: Expense side ──────────────────────────────────────────────
    # Entertainment excluded from both boxes per Art.53(1)(b)
    std_net = _amt(purch_std)                  # claimable only
    std_vat = std_net * 0.05

    entertainment_net = _amt(purch_std_entertainment)
    entertainment_vat = entertainment_net * 0.05

    # Box 6: Taxable expenses (standard-rated claimable + reverse-charge net)
    box6 = std_net + rc_net

    # Box 7: Input VAT = std claimable VAT + RC self-assessed input VAT (Art. 48)
    box7 = std_vat + rc_vat

    # Box 8: Net VAT payable (+) or refundable (-)
    box8 = box2 - box7

    return {
        "box1_standard_rated_supplies": round(box1, 2),
        "box2_vat_on_supplies":         round(box2, 2),
        "box3_zero_rated_supplies":     round(box3, 2),
        "box4_exempt_supplies":         round(box4, 2),
        "box5_total_taxable_supplies":  round(box5, 2),
        "box6_taxable_expenses":        round(box6, 2),
        "box7_vat_on_expenses":         round(box7, 2),
        "box8_vat_payable_or_refundable": round(box8, 2),
        # Extra metadata for UI tooltips
        "_rc_net_aed":               round(rc_net, 2),
        "_rc_vat_aed":               round(rc_vat, 2),
        "_std_net_aed":              round(std_net, 2),
        "_entertainment_blocked_net_aed": round(entertainment_net, 2),
        "_entertainment_blocked_vat_aed": round(entertainment_vat, 2),
    }


def generate_pdf_report(vat_return: VATReturn, company: Company, transactions: List[Transaction]) -> bytes:
    """Generate PDF report formatted like FTA VAT return form"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=1  # Center
    )
    story.append(Paragraph("UAE FEDERAL TAX AUTHORITY", title_style))
    story.append(Paragraph("VAT RETURN FORM", title_style))
    story.append(Spacer(1, 12))
    
    # Company Information
    company_style = ParagraphStyle(
        'CompanyInfo',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    story.append(Paragraph(f"<b>Company:</b> {company.name}", company_style))
    story.append(Paragraph(f"<b>TRN:</b> {company.trn or 'N/A'}", company_style))
    story.append(Paragraph(f"<b>Period:</b> {vat_return.period_start.strftime('%d/%m/%Y')} to {vat_return.period_end.strftime('%d/%m/%Y')}", company_style))
    story.append(Spacer(1, 12))
    
    # VAT Return Boxes Table
    data = [
        ['Box', 'Description', 'Amount (AED)'],
        ['Box 1', 'Standard Rated Supplies', f"{vat_return.box1_standard_rated_supplies:,.2f}"],
        ['Box 2', 'VAT on Supplies (5%)', f"{vat_return.box2_vat_on_supplies:,.2f}"],
        ['Box 3', 'Zero Rated Supplies', f"{vat_return.box3_zero_rated_supplies:,.2f}"],
        ['Box 4', 'Exempt Supplies', f"{vat_return.box4_exempt_supplies:,.2f}"],
        ['Box 5', 'Total Taxable Supplies', f"{vat_return.box5_total_taxable_supplies:,.2f}"],
        ['Box 6', 'Taxable Expenses', f"{vat_return.box6_taxable_expenses:,.2f}"],
        ['Box 7', 'VAT on Expenses (5%)', f"{vat_return.box7_vat_on_expenses:,.2f}"],
        ['Box 8', 'VAT Payable/Refundable', f"{vat_return.box8_vat_payable_or_refundable:,.2f}"],
    ]
    
    table = Table(data, colWidths=[30*mm, 100*mm, 50*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 12))
    
    # Highlight Box 8
    if vat_return.box8_vat_payable_or_refundable > 0:
        story.append(Paragraph(f"<b>VAT Payable to FTA: AED {vat_return.box8_vat_payable_or_refundable:,.2f}</b>", 
                              ParagraphStyle('Highlight', parent=styles['Normal'], fontSize=12, textColor=colors.red)))
    else:
        story.append(Paragraph(f"<b>VAT Refundable from FTA: AED {abs(vat_return.box8_vat_payable_or_refundable):,.2f}</b>", 
                              ParagraphStyle('Highlight', parent=styles['Normal'], fontSize=12, textColor=colors.green)))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def generate_excel_report(vat_return: VATReturn, company: Company, transactions: List[Transaction]) -> bytes:
    """Generate Excel report with multiple sheets"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: VAT Return Summary
    ws1 = wb.create_sheet("VAT Return Summary")
    ws1.append(["UAE FEDERAL TAX AUTHORITY - VAT RETURN"])
    ws1.append(["Company:", company.name])
    ws1.append(["TRN:", company.trn or "N/A"])
    ws1.append(["Period:", f"{vat_return.period_start.strftime('%d/%m/%Y')} to {vat_return.period_end.strftime('%d/%m/%Y')}"])
    ws1.append([])
    
    headers = ["Box", "Description", "Amount (AED)"]
    ws1.append(headers)
    
    data = [
        ["Box 1", "Standard Rated Supplies", vat_return.box1_standard_rated_supplies],
        ["Box 2", "VAT on Supplies (5%)", vat_return.box2_vat_on_supplies],
        ["Box 3", "Zero Rated Supplies", vat_return.box3_zero_rated_supplies],
        ["Box 4", "Exempt Supplies", vat_return.box4_exempt_supplies],
        ["Box 5", "Total Taxable Supplies", vat_return.box5_total_taxable_supplies],
        ["Box 6", "Taxable Expenses", vat_return.box6_taxable_expenses],
        ["Box 7", "VAT on Expenses (5%)", vat_return.box7_vat_on_expenses],
        ["Box 8", "VAT Payable/Refundable", vat_return.box8_vat_payable_or_refundable],
    ]
    
    for row in data:
        ws1.append(row)
    
    # Format Sheet 1
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws1[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row in ws1.iter_rows(min_row=6, max_row=13):
        for cell in row:
            if cell.column == 3:  # Amount column
                cell.number_format = '#,##0.00'
    
    # Sheet 2: Sales Transactions Detail
    ws2 = wb.create_sheet("Sales Transactions")
    sales = [
        t
        for t in transactions
        if _transaction_side(t) == "sale"
        and (t.vat_treatment or "") in ["standard_rated", "zero_rated", "exempt"]
    ]
    ws2.append(["Date", "Description", "Vendor/Customer", "Invoice Number", "Amount (AED)", "VAT Treatment", "VAT Amount (AED)"])
    
    for t in sales:
        ws2.append([
            t.date.strftime('%d/%m/%Y'),
            t.description,
            t.vendor_or_customer or "",
            t.invoice_number or "",
            t.amount_aed,
            t.vat_treatment or "",
            t.vat_amount_aed
        ])
    
    # Format Sheet 2
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 3: Purchase Transactions Detail
    ws3 = wb.create_sheet("Purchase Transactions")
    purchases = [
        t
        for t in transactions
        if _transaction_side(t) == "purchase"
        and (t.vat_treatment or "") == "standard_rated"
        and t.amount_aed > 0
    ]
    ws3.append(["Date", "Description", "Vendor/Customer", "Invoice Number", "Amount (AED)", "VAT Amount (AED)"])
    
    for t in purchases:
        ws3.append([
            t.date.strftime('%d/%m/%Y'),
            t.description,
            t.vendor_or_customer or "",
            t.invoice_number or "",
            t.amount_aed,
            t.vat_amount_aed
        ])
    
    # Format Sheet 3
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 4: Zero-rated breakdown
    ws4 = wb.create_sheet("Zero Rated Breakdown")
    zero_rated = [
        t for t in transactions if _transaction_side(t) == "sale" and t.vat_treatment == "zero_rated"
    ]
    ws4.append(["Date", "Description", "Vendor/Customer", "Invoice Number", "Amount (AED)"])
    
    for t in zero_rated:
        ws4.append([
            t.date.strftime('%d/%m/%Y'),
            t.description,
            t.vendor_or_customer or "",
            t.invoice_number or "",
            t.amount_aed
        ])
    
    # Format Sheet 4
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 5: Exempt breakdown
    ws5 = wb.create_sheet("Exempt Breakdown")
    exempt = [t for t in transactions if _transaction_side(t) == "sale" and t.vat_treatment == "exempt"]
    ws5.append(["Date", "Description", "Vendor/Customer", "Invoice Number", "Amount (AED)"])
    
    for t in exempt:
        ws5.append([
            t.date.strftime('%d/%m/%Y'),
            t.description,
            t.vendor_or_customer or "",
            t.invoice_number or "",
            t.amount_aed
        ])
    
    # Format Sheet 5
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.post("/generate-return", response_model=VATReturnResponse)
async def generate_return(
    request: GenerateReturnRequest,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """
    Generate VAT return for a company and period.
    
    Pulls all verified transactions, calculates 8 FTA boxes,
    saves to database, and generates PDF/Excel reports.
    """
    # Use auth-verified company_id, not request body
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # Get verified transactions for the period
    transactions = db.query(Transaction).filter(
        and_(
            Transaction.company_id == company_id,
            Transaction.date >= request.period_start,
            Transaction.date <= request.period_end,
            Transaction.is_verified == True
        )
    ).all()
    
    if not transactions:
        raise HTTPException(status_code=400, detail="No verified transactions found for this period")
    
    # Calculate VAT return boxes
    box_values = calculate_vat_return_boxes(transactions)
    
    # Check if return already exists
    existing_return = db.query(VATReturn).filter(
        and_(
            VATReturn.company_id == company_id,
            VATReturn.period_start == request.period_start,
            VATReturn.period_end == request.period_end
        )
    ).first()
    
    if existing_return:
        # Update existing return
        for key, value in box_values.items():
            setattr(existing_return, key, value)
        existing_return.status = "draft"
        vat_return = existing_return
    else:
        # Create new return
        vat_return = VATReturn(
            company_id=company_id,
            period_start=request.period_start,
            period_end=request.period_end,
            **box_values,
            status="draft"
        )
        db.add(vat_return)
    
    db.commit()
    db.refresh(vat_return)
    
    # Generate PDF and Excel
    pdf_content = generate_pdf_report(vat_return, company, transactions)
    excel_content = generate_excel_report(vat_return, company, transactions)
    
    # Save files (in production, save to S3 or file storage)
    # For now, we'll return them as downloadable
    pdf_filename = f"vat_return_{company.id}_{vat_return.id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    excel_filename = f"vat_return_{company.id}_{vat_return.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return {
        "return_id": vat_return.id,
        "company_id": vat_return.company_id,
        "period_start": vat_return.period_start,
        "period_end": vat_return.period_end,
        **box_values,
        "status": vat_return.status,
        "created_at": vat_return.created_at,
        "pdf_url": f"/api/vat/returns/{vat_return.id}/pdf",
        "excel_url": f"/api/vat/returns/{vat_return.id}/excel",
    }


@router.get("/returns/{return_id}/pdf")
async def get_return_pdf(
    return_id: int,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """Download VAT return PDF"""
    vat_return = db.query(VATReturn).filter(
        VATReturn.id == return_id, VATReturn.company_id == company_id
    ).first()
    if not vat_return:
        raise HTTPException(status_code=404, detail="VAT return not found")

    company = db.query(Company).filter(Company.id == company_id).first()
    transactions = db.query(Transaction).filter(
        and_(
            Transaction.company_id == company_id,
            Transaction.date >= vat_return.period_start,
            Transaction.date <= vat_return.period_end,
            Transaction.is_verified == True,
        )
    ).all()

    pdf_content = generate_pdf_report(vat_return, company, transactions)

    return StreamingResponse(
        io.BytesIO(pdf_content),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=vat_return_{return_id}.pdf"},
    )


@router.get("/returns/{return_id}/excel")
async def get_return_excel(
    return_id: int,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """Download VAT return Excel"""
    vat_return = db.query(VATReturn).filter(
        VATReturn.id == return_id, VATReturn.company_id == company_id
    ).first()
    if not vat_return:
        raise HTTPException(status_code=404, detail="VAT return not found")

    company = db.query(Company).filter(Company.id == company_id).first()
    transactions = db.query(Transaction).filter(
        and_(
            Transaction.company_id == company_id,
            Transaction.date >= vat_return.period_start,
            Transaction.date <= vat_return.period_end,
            Transaction.is_verified == True,
        )
    ).all()

    excel_content = generate_excel_report(vat_return, company, transactions)
    
    return StreamingResponse(
        io.BytesIO(excel_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vat_return_{return_id}.xlsx"}
    )


@router.post("/returns/{return_id}/submit", response_model=VATSubmitResponse)
async def submit_vat_return(
    return_id: int,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """
    FTA submission tracked stub.

    For now this route records an attempted submission and marks the return as failed
    with a stubbed integration error, so frontend flows can handle a stable contract.
    """
    vat_return = db.query(VATReturn).filter(
        VATReturn.id == return_id, VATReturn.company_id == company_id
    ).first()
    if not vat_return:
        raise HTTPException(status_code=404, detail="VAT return not found")

    payload_snapshot = {
        "return_id": vat_return.id,
        "company_id": vat_return.company_id,
        "period_start": vat_return.period_start.isoformat(),
        "period_end": vat_return.period_end.isoformat(),
        "box1_standard_rated_supplies": vat_return.box1_standard_rated_supplies,
        "box2_vat_on_supplies": vat_return.box2_vat_on_supplies,
        "box3_zero_rated_supplies": vat_return.box3_zero_rated_supplies,
        "box4_exempt_supplies": vat_return.box4_exempt_supplies,
        "box5_total_taxable_supplies": vat_return.box5_total_taxable_supplies,
        "box6_taxable_expenses": vat_return.box6_taxable_expenses,
        "box7_vat_on_expenses": vat_return.box7_vat_on_expenses,
        "box8_vat_payable_or_refundable": vat_return.box8_vat_payable_or_refundable,
    }

    stub_error = "FTA submit integration is not configured"
    attempted_at = datetime.utcnow()
    vat_return.submission_status = "failed"
    vat_return.submitted_at = attempted_at
    vat_return.fta_reference_number = None
    vat_return.submission_error = stub_error

    db.add(
        FTASubmissionLog(
            vat_return_id=vat_return.id,
            attempted_at=attempted_at,
            submission_status=vat_return.submission_status,
            payload_snapshot=payload_snapshot,
            response_raw={"stub": True, "error": stub_error},
        )
    )
    db.commit()
    db.refresh(vat_return)

    return {
        "return_id": vat_return.id,
        "submission_status": vat_return.submission_status,
        "submitted_at": vat_return.submitted_at,
        "fta_reference_number": vat_return.fta_reference_number,
        "submission_error": vat_return.submission_error,
    }


class VATAssessedPayload(BaseModel):
    """Inbound VAT assessment payload from n8n UAE_VAT_Return_Prep workflow."""
    company_id: int
    period: str  # e.g. "2024-Q1", "2024-01", "2024"
    fiscal_year: Optional[str] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    box1_standard_rated_supplies: float = 0.0
    box2_vat_on_supplies: float = 0.0
    box3_zero_rated_supplies: float = 0.0
    box4_exempt_supplies: float = 0.0
    box5_total_taxable_supplies: float = 0.0
    box6_taxable_expenses: float = 0.0
    box7_vat_on_expenses: float = 0.0
    box8_vat_payable_or_refundable: float = 0.0
    box9: Optional[float] = None
    box10: Optional[float] = None
    box11: Optional[float] = None
    filing_deadline: Optional[str] = None
    days_to_filing: Optional[int] = None
    is_overdue: Optional[bool] = None
    urgency_level: Optional[str] = None
    assessment_date: Optional[str] = None


def _verify_vat_signature(secret: str, body: bytes, received: str) -> bool:
    """HMAC-SHA256 hex over raw body, or plain shared secret fallback."""
    if not received or not received.strip():
        return False
    received = received.strip()
    expected_hex = hmac.new(secret.encode("utf-8"), body, hashlib.sha256).hexdigest()
    if hmac.compare_digest(expected_hex, received.lower()):
        return True
    return secrets.compare_digest(received, secret)


def _period_to_dates(period: str, period_start: Optional[date], period_end: Optional[date]):
    """Resolve period_start / period_end from explicit dates or period string."""
    if period_start and period_end:
        return period_start, period_end
    p = period.strip()
    # YYYY-QN
    if len(p) == 7 and p[4] == "-" and p[5] == "Q":
        year = int(p[:4])
        quarter = int(p[6])
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        import calendar
        ps = date(year, start_month, 1)
        pe = date(year, end_month, calendar.monthrange(year, end_month)[1])
        return ps, pe
    # YYYY-MM
    if len(p) == 7 and p[4] == "-":
        import calendar
        year, month = int(p[:4]), int(p[5:])
        ps = date(year, month, 1)
        pe = date(year, month, calendar.monthrange(year, month)[1])
        return ps, pe
    # YYYY
    if len(p) == 4 and p.isdigit():
        year = int(p)
        return date(year, 1, 1), date(year, 12, 31)
    raise ValueError(f"Cannot parse period '{period}' — use YYYY-QN, YYYY-MM, or YYYY")


@router.post("/n8n/inbound/vat_assessed")
async def inbound_vat_assessed(
    request: Request,
    db: Session = Depends(get_db),
    x_n8n_signature: Optional[str] = Header(default=None, alias="X-N8N-Signature"),
):
    """Receive signed VAT assessment from n8n and upsert vat_returns by company + period."""
    secret = os.getenv("N8N_WEBHOOK_SECRET")
    if not secret:
        raise HTTPException(status_code=500, detail="N8N_WEBHOOK_SECRET is not configured")
    if not x_n8n_signature:
        raise HTTPException(status_code=401, detail="Missing X-N8N-Signature header")

    raw_body = await request.body()
    if not _verify_vat_signature(secret=secret, body=raw_body, received=x_n8n_signature):
        raise HTTPException(status_code=401, detail="Invalid signature")

    try:
        payload = VATAssessedPayload.model_validate_json(raw_body)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors()) from exc

    company = db.query(Company).filter(Company.id == payload.company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    try:
        ps, pe = _period_to_dates(payload.period, payload.period_start, payload.period_end)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    existing = db.query(VATReturn).filter(
        and_(
            VATReturn.company_id == payload.company_id,
            VATReturn.period_start == ps,
        )
    ).first()

    box_values = {
        "box1_standard_rated_supplies": payload.box1_standard_rated_supplies,
        "box2_vat_on_supplies": payload.box2_vat_on_supplies,
        "box3_zero_rated_supplies": payload.box3_zero_rated_supplies,
        "box4_exempt_supplies": payload.box4_exempt_supplies,
        "box5_total_taxable_supplies": payload.box5_total_taxable_supplies,
        "box6_taxable_expenses": payload.box6_taxable_expenses,
        "box7_vat_on_expenses": payload.box7_vat_on_expenses,
        "box8_vat_payable_or_refundable": payload.box8_vat_payable_or_refundable,
    }

    if existing:
        for k, v in box_values.items():
            setattr(existing, k, v)
        vat_return = existing
    else:
        vat_return = VATReturn(
            company_id=payload.company_id,
            period_start=ps,
            period_end=pe,
            status="draft",
            **box_values,
        )
        db.add(vat_return)

    db.flush()

    db.add(AuditLog(
        company_id=payload.company_id,
        actor="n8n",
        entity_type="vat_return",
        entity_id=vat_return.id,
        action="assessed",
        after_state={
            "period": payload.period,
            "box8": payload.box8_vat_payable_or_refundable,
            "urgency_level": payload.urgency_level,
            "is_overdue": payload.is_overdue,
        },
    ))

    db.commit()
    db.refresh(vat_return)
    return {"status": "ok", "return_id": vat_return.id}


@router.get("/returns")
async def list_vat_returns(
    company_id: int = Depends(get_current_company_id),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """List VAT returns for a company."""
    rows = (
        db.query(VATReturn)
        .filter(VATReturn.company_id == company_id)
        .order_by(VATReturn.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "return_id": r.id,
            "company_id": r.company_id,
            "period_start": r.period_start,
            "period_end": r.period_end,
            "box8_vat_payable_or_refundable": r.box8_vat_payable_or_refundable,
            "status": r.status,
            "submission_status": r.submission_status,
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/reconcile/{vat_return_id}", response_model=ReconciliationResponse)
async def reconcile_return(
    vat_return_id: int,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """
    Reconcile VAT return with invoice transactions.
    
    Compares invoice totals by treatment against VAT return boxes,
    finds mismatches > AED 100, and provides AI recommendations.
    """
    # Get VAT return — must belong to the verified company
    vat_return = db.query(VATReturn).filter(
        VATReturn.id == vat_return_id, VATReturn.company_id == company_id
    ).first()
    if not vat_return:
        raise HTTPException(status_code=404, detail="VAT return not found")

    # Get all transactions for the period
    transactions = db.query(Transaction).filter(
        and_(
            Transaction.company_id == company_id,
            Transaction.date >= vat_return.period_start,
            Transaction.date <= vat_return.period_end
        )
    ).all()
    
    # Sum invoice amounts by treatment
    invoice_totals = {}
    invoice_vat_totals = {}
    
    for t in transactions:
        treatment = t.vat_treatment or "unclassified"
        if treatment not in invoice_totals:
            invoice_totals[treatment] = 0.0
            invoice_vat_totals[treatment] = 0.0
        invoice_totals[treatment] += t.amount_aed
        invoice_vat_totals[treatment] += t.vat_amount_aed
    
    # Compare against VAT return boxes
    mismatches = []
    total_difference = 0.0
    
    # Box 1: Standard rated supplies
    invoice_standard = invoice_totals.get("standard_rated", 0.0)
    difference_1 = abs(invoice_standard - vat_return.box1_standard_rated_supplies)
    if difference_1 > 100:
        mismatches.append({
            "box": "Box 1",
            "issue": "Standard rated supplies amount mismatch",
            "invoice_amount": invoice_standard,
            "return_amount": vat_return.box1_standard_rated_supplies,
            "difference": difference_1
        })
        total_difference += difference_1
    
    # Box 2: Output VAT
    invoice_output_vat = invoice_vat_totals.get("standard_rated", 0.0)
    difference_2 = abs(invoice_output_vat - vat_return.box2_vat_on_supplies)
    if difference_2 > 100:
        mismatches.append({
            "box": "Box 2",
            "issue": "Output VAT mismatch",
            "invoice_amount": invoice_output_vat,
            "return_amount": vat_return.box2_vat_on_supplies,
            "difference": difference_2
        })
        total_difference += difference_2
    
    # Box 3: Zero rated
    invoice_zero = invoice_totals.get("zero_rated", 0.0)
    difference_3 = abs(invoice_zero - vat_return.box3_zero_rated_supplies)
    if difference_3 > 100:
        mismatches.append({
            "box": "Box 3",
            "issue": "Zero rated supplies mismatch",
            "invoice_amount": invoice_zero,
            "return_amount": vat_return.box3_zero_rated_supplies,
            "difference": difference_3
        })
        total_difference += difference_3
    
    # Box 4: Exempt
    invoice_exempt = invoice_totals.get("exempt", 0.0)
    difference_4 = abs(invoice_exempt - vat_return.box4_exempt_supplies)
    if difference_4 > 100:
        mismatches.append({
            "box": "Box 4",
            "issue": "Exempt supplies mismatch",
            "invoice_amount": invoice_exempt,
            "return_amount": vat_return.box4_exempt_supplies,
            "difference": difference_4
        })
        total_difference += difference_4
    
    # Find specific transaction mismatches
    transaction_mismatches = []
    for t in transactions:
        if t.invoice_number:
            # Check if transaction amount matches expected VAT treatment
            expected_vat = 0.0
            if t.vat_treatment == "standard_rated":
                expected_vat = t.amount_aed * 0.05
            elif t.vat_treatment == "zero_rated":
                expected_vat = 0.0
            elif t.vat_treatment == "exempt":
                expected_vat = 0.0
            
            vat_diff = abs(t.vat_amount_aed - expected_vat)
            if vat_diff > 100:
                transaction_mismatches.append({
                    "invoice_number": t.invoice_number,
                    "issue": f"VAT treatment mismatch or amount difference",
                    "transaction_amount": t.amount_aed,
                    "return_amount": expected_vat,
                    "difference": vat_diff
                })
    
    # Format mismatches to match requested format
    all_mismatches = []
    for m in mismatches:
        all_mismatches.append({
            "invoice_number": m.get("box", "N/A"),
            "issue": m["issue"],
            "transaction_amount": m["invoice_amount"],
            "return_amount": m["return_amount"],
            "difference": m["difference"]
        })
    
    all_mismatches.extend(transaction_mismatches)
    
    # Generate AI recommendation
    recommendation = "No issues found. VAT return matches invoice totals."
    if all_mismatches:
        if claude_client:
            try:
                mismatch_summary = "\n".join([
                    f"- {m.get('issue', 'Unknown issue')}: Difference of AED {m.get('difference', 0):,.2f}"
                    for m in all_mismatches[:10]  # Limit to first 10
                ])
                
                prompt = f"""You are a UAE VAT expert. A VAT return reconciliation found the following mismatches:

{mismatch_summary}

Provide a brief recommendation (2-3 sentences) on how to fix these issues. Focus on common causes like:
- Unverified transactions
- Incorrect VAT treatment classification
- Missing invoices
- Calculation errors

Return only the recommendation text, no markdown or formatting."""
                
                message = claude_client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=200,
                    temperature=0.3,
                    messages=[{"role": "user", "content": prompt}]
                )
                recommendation = message.content[0].text.strip()
            except Exception as e:
                recommendation = f"Review mismatches manually. Common causes: unverified transactions, incorrect classifications, or missing invoices. Error: {str(e)}"
        else:
            recommendation = "Review mismatches manually. Common causes: unverified transactions, incorrect VAT treatment classifications, or missing invoices."
    
    # Save reconciliation result
    reconciliation = ReconciliationResult(
        company_id=company_id,
        vat_return_id=vat_return_id,
        total_invoices_aed=sum(invoice_totals.values()),
        total_output_vat_aed=invoice_output_vat,
        vat_return_output_aed=vat_return.box2_vat_on_supplies,
        difference_aed=total_difference,
        mismatches=all_mismatches,
        status="matched" if not all_mismatches else "mismatch_found"
    )
    
    db.add(reconciliation)
    db.commit()
    
    return {
        "status": reconciliation.status,
        "difference_aed": total_difference,
        "mismatches": all_mismatches,
        "recommendation": recommendation
    }
