"""
Generate a test Excel file with 20 UAE VAT transactions covering all treatment types.
Run: python create_test_transactions.py
Output: UAE_VAT_Test_Transactions.xlsx (in current directory)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Data ──────────────────────────────────────────────────────────────────────

TRANSACTIONS = [
    # Standard Rated Sales (5% VAT output)
    {
        "date": "2025-01-05",
        "type": "sale",
        "description": "Consulting services to ABC Corp (standard rated)",
        "supplier_customer": "ABC Corp LLC",
        "trn": "100345678900003",
        "invoice_no": "INV-2025-001",
        "amount_aed": 50000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 2500.00,
        "vat_treatment": "standard_rated",
        "notes": "B2B service — VAT applies at 5%",
    },
    {
        "date": "2025-01-12",
        "type": "sale",
        "description": "Software licence sale to XYZ Trading",
        "supplier_customer": "XYZ Trading FZE",
        "trn": "100987654300001",
        "invoice_no": "INV-2025-002",
        "amount_aed": 120000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 6000.00,
        "vat_treatment": "standard_rated",
        "notes": "Annual licence — domestic supply",
    },
    {
        "date": "2025-01-20",
        "type": "sale",
        "description": "Training course fees — Dubai branch",
        "supplier_customer": "Gulf Industries LLC",
        "trn": "100112233440003",
        "invoice_no": "INV-2025-003",
        "amount_aed": 25000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 1250.00,
        "vat_treatment": "standard_rated",
        "notes": "Educational service taxable at 5%",
    },
    {
        "date": "2025-02-03",
        "type": "sale",
        "description": "IT equipment sale — servers and racks",
        "supplier_customer": "Emirates Bank PJSC",
        "trn": "100556677880001",
        "invoice_no": "INV-2025-004",
        "amount_aed": 380000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 19000.00,
        "vat_treatment": "standard_rated",
        "notes": "Capital goods sale — standard rated",
    },
    {
        "date": "2025-02-15",
        "type": "sale",
        "description": "Advisory retainer Q1 — Sharjah client",
        "supplier_customer": "Sharjah Holdings Ltd",
        "trn": "100334455660002",
        "invoice_no": "INV-2025-005",
        "amount_aed": 75000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 3750.00,
        "vat_treatment": "standard_rated",
        "notes": "Monthly retainer — standard rated",
    },
    # Zero Rated Supplies
    {
        "date": "2025-01-08",
        "type": "sale",
        "description": "Export of goods to Saudi Arabia — FOB Jebel Ali",
        "supplier_customer": "Saudi Exports Co",
        "trn": "",
        "invoice_no": "INV-2025-006",
        "amount_aed": 200000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "zero_rated",
        "notes": "Art.30 UAE VAT Law — export of goods zero rated. Retain export documents.",
    },
    {
        "date": "2025-02-10",
        "type": "sale",
        "description": "International freight services — sea cargo",
        "supplier_customer": "Global Freight DMCC",
        "trn": "100778899000001",
        "invoice_no": "INV-2025-007",
        "amount_aed": 45000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "zero_rated",
        "notes": "Zero rated international transport Art.33",
    },
    {
        "date": "2025-03-01",
        "type": "sale",
        "description": "Basic food supply — dates and grains",
        "supplier_customer": "Al Baraka Supermarket",
        "trn": "100223344550004",
        "invoice_no": "INV-2025-008",
        "amount_aed": 18000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "zero_rated",
        "notes": "Zero rated basic food items — Schedule 2 UAE VAT Decree",
    },
    # Exempt Supplies
    {
        "date": "2025-01-25",
        "type": "sale",
        "description": "Residential property lease — Dubai apartment",
        "supplier_customer": "Al Waha Properties",
        "trn": "",
        "invoice_no": "LEASE-2025-001",
        "amount_aed": 90000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "exempt",
        "notes": "Exempt — residential lease Art.46(2). Input tax blocked on related costs.",
    },
    {
        "date": "2025-02-28",
        "type": "sale",
        "description": "Life insurance premium collected",
        "supplier_customer": "Individual Client",
        "trn": "",
        "invoice_no": "INS-2025-001",
        "amount_aed": 15000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "exempt",
        "notes": "Exempt financial service — life insurance Art.46(4)",
    },
    # Reverse Charge (Purchases from non-UAE suppliers)
    {
        "date": "2025-01-15",
        "type": "purchase",
        "description": "Cloud hosting services — AWS Ireland",
        "supplier_customer": "Amazon Web Services EMEA",
        "trn": "",
        "invoice_no": "AWS-JAN-2025",
        "amount_aed": 35000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 1750.00,
        "vat_treatment": "reverse_charge",
        "notes": "Imported service — buyer accounts for VAT under Art.48. Declare in Box 3 & Box 9.",
    },
    {
        "date": "2025-02-05",
        "type": "purchase",
        "description": "Microsoft 365 enterprise licences — Ireland",
        "supplier_customer": "Microsoft Ireland Operations",
        "trn": "",
        "invoice_no": "MSFT-FEB25",
        "amount_aed": 22000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 1100.00,
        "vat_treatment": "reverse_charge",
        "notes": "Imported service — reverse charge applies. Self-assess 5% VAT.",
    },
    {
        "date": "2025-03-10",
        "type": "purchase",
        "description": "Legal advisory fees — UK law firm",
        "supplier_customer": "Clifford Chance LLP",
        "trn": "",
        "invoice_no": "CC-MAR25-001",
        "amount_aed": 85000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 4250.00,
        "vat_treatment": "reverse_charge",
        "notes": "Cross-border professional service — reverse charge Art.48. Substantive input tax claim.",
    },
    # Standard Rated Purchases (Input Tax)
    {
        "date": "2025-01-10",
        "type": "purchase",
        "description": "Office rent — DIFC Tower 2 Q1",
        "supplier_customer": "DIFC Investments LLC",
        "trn": "100445566770001",
        "invoice_no": "DIFC-2025-Q1",
        "amount_aed": 62500.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 3125.00,
        "vat_treatment": "standard_rated",
        "notes": "Commercial rent — input tax fully reclaimable",
    },
    {
        "date": "2025-01-22",
        "type": "purchase",
        "description": "Laptop computers — 10 units",
        "supplier_customer": "Jumbo Electronics LLC",
        "trn": "100667788990002",
        "invoice_no": "JMB-2025-0122",
        "amount_aed": 55000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 2750.00,
        "vat_treatment": "standard_rated",
        "notes": "Capital equipment — input tax reclaimable if used for taxable supply",
    },
    {
        "date": "2025-02-14",
        "type": "purchase",
        "description": "Marketing agency retainer — Feb 2025",
        "supplier_customer": "Creative Hub DMCC",
        "trn": "100889900110003",
        "invoice_no": "CH-FEB25",
        "amount_aed": 30000.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 1500.00,
        "vat_treatment": "standard_rated",
        "notes": "Business overhead — fully recoverable input tax",
    },
    {
        "date": "2025-02-20",
        "type": "purchase",
        "description": "Staff entertainment — team dinner",
        "supplier_customer": "Nobu Restaurant DIFC",
        "trn": "100112200330001",
        "invoice_no": "NOBU-2025-0220",
        "amount_aed": 8500.00,
        "vat_rate_pct": 5,
        "vat_amount_aed": 425.00,
        "vat_treatment": "standard_rated",
        "notes": "BLOCKED INPUT TAX — entertainment expenses Art.53. VAT not recoverable.",
    },
    # Out of Scope
    {
        "date": "2025-01-31",
        "type": "sale",
        "description": "Dividend income from subsidiary",
        "supplier_customer": "GulfTax Holdings Ltd",
        "trn": "",
        "invoice_no": "DIV-2025-001",
        "amount_aed": 500000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "out_of_scope",
        "notes": "Dividend — outside scope of UAE VAT. No impact on VAT return.",
    },
    {
        "date": "2025-02-25",
        "type": "purchase",
        "description": "Salary payments — February 2025",
        "supplier_customer": "Employees (Payroll)",
        "trn": "",
        "invoice_no": "PAYROLL-FEB25",
        "amount_aed": 280000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "out_of_scope",
        "notes": "Employment income — outside scope of VAT. No VAT treatment required.",
    },
    {
        "date": "2025-03-15",
        "type": "sale",
        "description": "Intercompany loan interest charged",
        "supplier_customer": "GulfTax FZ LLC",
        "trn": "",
        "invoice_no": "IC-INT-Q1",
        "amount_aed": 12000.00,
        "vat_rate_pct": 0,
        "vat_amount_aed": 0.00,
        "vat_treatment": "out_of_scope",
        "notes": "Intercompany interest — outside scope. Check Art.1 UAE VAT Law definition of supply.",
    },
]

# ── Excel Build ───────────────────────────────────────────────────────────────

def build_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VAT Transactions Q1-2025"

    # Colors
    HEADER_BG = "1E3A5F"
    HEADER_FG = "FFFFFF"
    STANDARD_BG = "E8F4FD"
    ZERO_BG = "E8F8E8"
    EXEMPT_BG = "FFF8E1"
    RC_BG = "FDE8F0"
    OOS_BG = "F5F5F5"

    VAT_COLORS = {
        "standard_rated": STANDARD_BG,
        "zero_rated": ZERO_BG,
        "exempt": EXEMPT_BG,
        "reverse_charge": RC_BG,
        "out_of_scope": OOS_BG,
    }

    TREATMENT_LABELS = {
        "standard_rated": "Standard Rated (5%)",
        "zero_rated": "Zero Rated (0%)",
        "exempt": "Exempt",
        "reverse_charge": "Reverse Charge",
        "out_of_scope": "Out of Scope",
    }

    # ── Title block ──
    ws.merge_cells("A1:N1")
    ws["A1"] = "GulfTax AI — UAE VAT Test Transactions — Q1 2025"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = "Federal Decree-Law No.8 of 2017 on Value Added Tax | Import to GulfTax AI at /dashboard/vat-classifier"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[3].height = 8  # spacer

    # ── Column headers ──
    HEADERS = [
        "Date", "Type", "Description", "Supplier / Customer",
        "TRN", "Invoice No", "Amount AED", "VAT Rate %",
        "VAT Amount AED", "Total AED", "VAT Treatment",
        "Treatment Label", "Notes / Compliance Flag",
    ]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font = Font(bold=True, color=HEADER_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[4].height = 30

    # ── Data rows ──
    for row_idx, txn in enumerate(TRANSACTIONS, start=5):
        total = txn["amount_aed"] + txn["vat_amount_aed"]
        values = [
            txn["date"],
            txn["type"].upper(),
            txn["description"],
            txn["supplier_customer"],
            txn["trn"] or "N/A",
            txn["invoice_no"],
            txn["amount_aed"],
            txn["vat_rate_pct"],
            txn["vat_amount_aed"],
            total,
            txn["vat_treatment"],
            TREATMENT_LABELS.get(txn["vat_treatment"], txn["vat_treatment"]),
            txn["notes"],
        ]

        bg = VAT_COLORS.get(txn["vat_treatment"], "FFFFFF")
        fill = PatternFill("solid", fgColor=bg)

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 13))
            cell.font = Font(size=9)

            # Format currency columns
            if col in (7, 9, 10):
                cell.number_format = '#,##0.00'
            if col == 8:
                cell.number_format = '0"%"'

        ws.row_dimensions[row_idx].height = 40

    # ── Summary block ──
    sum_row = len(TRANSACTIONS) + 7

    def sum_section(label, treatment):
        txns_filtered = [t for t in TRANSACTIONS if t["vat_treatment"] == treatment]
        return {
            "label": label,
            "count": len(txns_filtered),
            "amount": sum(t["amount_aed"] for t in txns_filtered),
            "vat": sum(t["vat_amount_aed"] for t in txns_filtered),
        }

    summaries = [
        sum_section("Standard Rated Supplies", "standard_rated"),
        sum_section("Zero Rated Supplies", "zero_rated"),
        sum_section("Exempt Supplies", "exempt"),
        sum_section("Reverse Charge", "reverse_charge"),
        sum_section("Out of Scope", "out_of_scope"),
    ]

    ws.cell(row=sum_row, column=1, value="SUMMARY BY VAT TREATMENT").font = Font(bold=True, size=11, color="1E3A5F")

    hdr_row = sum_row + 1
    for col, txt in enumerate(["Treatment", "Count", "Net Amount AED", "VAT AED"], start=1):
        c = ws.cell(row=hdr_row, column=col, value=txt)
        c.font = Font(bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.border = border
        c.alignment = Alignment(horizontal="center")

    for i, s in enumerate(summaries):
        r = hdr_row + 1 + i
        bg = list(VAT_COLORS.values())[i]
        fill = PatternFill("solid", fgColor=bg)
        for col, val in enumerate([s["label"], s["count"], s["amount"], s["vat"]], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.border = border
            c.font = Font(size=9, bold=(col == 1))
            if col in (3, 4):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")

    # Grand total
    total_r = hdr_row + len(summaries) + 1
    ws.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_r, column=3, value=sum(t["amount_aed"] for t in TRANSACTIONS)).number_format = '#,##0.00'
    ws.cell(row=total_r, column=4, value=sum(t["vat_amount_aed"] for t in TRANSACTIONS)).number_format = '#,##0.00'
    for col in range(1, 5):
        ws.cell(row=total_r, column=col).font = Font(bold=True)
        ws.cell(row=total_r, column=col).fill = PatternFill("solid", fgColor="E0E8F0")
        ws.cell(row=total_r, column=col).border = border

    # ── Column widths ──
    WIDTHS = [12, 8, 40, 28, 18, 16, 14, 10, 14, 14, 18, 22, 55]
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = "A5"

    # ── Legend sheet ──
    ls = wb.create_sheet("VAT Treatment Guide")
    ls["A1"] = "UAE VAT Treatment Reference Guide"
    ls["A1"].font = Font(bold=True, size=13, color="1E3A5F")

    guide = [
        ("Treatment", "Rate", "UAE Law Reference", "Key Rule"),
        ("Standard Rated", "5%", "Art.3, Federal Decree-Law No.8/2017", "Default rate for all taxable supplies within UAE"),
        ("Zero Rated", "0%", "Art.30-45, Federal Decree-Law No.8/2017", "Taxable at 0%; input tax fully reclaimable. Includes exports, international transport, first supply of new residential buildings"),
        ("Exempt", "N/A", "Art.46, Federal Decree-Law No.8/2017", "No VAT charged; input tax on related costs BLOCKED. Includes residential leases, financial services, bare land"),
        ("Reverse Charge", "5% (self)", "Art.48, Federal Decree-Law No.8/2017", "Imported services — buyer accounts for VAT. Declare both output (Box 3) and input (Box 9) in VAT return"),
        ("Out of Scope", "N/A", "Art.1 — definition of 'Supply'", "Not a supply under UAE VAT law. No VAT return impact. Includes salaries, dividends, intercompany transfers"),
    ]

    for r, row_data in enumerate(guide, start=3):
        for c, val in enumerate(row_data, start=1):
            cell = ls.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ls.row_dimensions[r].height = 50 if r > 3 else 20

    ls.column_dimensions["A"].width = 20
    ls.column_dimensions["B"].width = 10
    ls.column_dimensions["C"].width = 35
    ls.column_dimensions["D"].width = 60

    # ── Save ──
    output_path = "UAE_VAT_Test_Transactions_Q1_2025.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"   {len(TRANSACTIONS)} transactions across 5 VAT treatment types")
    print(f"   Total net: AED {sum(t['amount_aed'] for t in TRANSACTIONS):,.2f}")
    print(f"   Total VAT: AED {sum(t['vat_amount_aed'] for t in TRANSACTIONS):,.2f}")

if __name__ == "__main__":
    build_excel()
