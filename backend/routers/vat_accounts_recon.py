"""VAT Reconciliation vs Accounts — GL tie-out to VAT return."""
from __future__ import annotations

import base64
import json
import logging
import os
import re
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from xml.sax.saxutils import escape

import pandas as pd
from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from database import get_db
from middleware.auth import get_current_company_id
from models import Company, VATAccountsReconciliation, VATReturn

load_dotenv()
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/vat-accounts", tags=["VAT vs Accounts"])

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 10_000
RECON_TOLERANCE = 1.0

anthropic_api_key = os.getenv("ANTHROPIC_API_KEY")
claude_client = Anthropic(api_key=anthropic_api_key) if anthropic_api_key else None

_ENTERTAINMENT_KWS = ("nobu", "restaurant", "entertainment", "meals", "hotel", "cafe")
_BLOCKED_KWS = ("blocked", "non-claimable", "non claimable", "irrecoverable")


def _get_supabase():
    url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
    key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_KEY") or "").strip()
    if not url or not key:
        return None
    try:
        from supabase import create_client  # type: ignore
        return create_client(url, key)
    except Exception as exc:
        logger.warning("Supabase client init failed: %s", exc)
        return None


def _json_from_claude(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    if "```json" in cleaned:
        cleaned = cleaned.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in cleaned:
        cleaned = cleaned.split("```", 1)[1].split("```", 1)[0].strip()
    return json.loads(cleaned)


def _to_float(val: Any) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("AED", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fmt_aed(amount: float) -> str:
    return f"AED {amount:,.2f}"


def _pdf_text(value: Any) -> str:
    return escape(str(value or ""))


def _pick_column(columns: List[str], patterns: Tuple[str, ...]) -> Optional[str]:
    for col in columns:
        cl = col.lower().strip()
        for pat in patterns:
            if pat == cl or pat in cl:
                return col
    return None


def _detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    cols = [str(c).strip().lower() for c in df.columns]
    df.columns = cols
    mapping: Dict[str, Optional[str]] = {
        "date": _pick_column(cols, ("date", "transaction date", "voucher date", "posting date")),
        "voucher": _pick_column(cols, ("voucher no", "voucher", "ref", "reference", "invoice no", "no.", "no")),
        "party": _pick_column(cols, ("party", "vendor", "customer", "supplier", "name")),
        "trn": _pick_column(cols, ("trn", "tax number", "vat number", "party tax id", "tax registration")),
        "amount": _pick_column(cols, ("taxable amount", "net amount", "amount", "debit", "credit")),
        "vat_amount": _pick_column(cols, ("vat amount", "tax amount", "input vat", "output vat", "tax value", "tax")),
        "type": _pick_column(cols, ("type", "transaction type", "classification", "tax type")),
        "account": _pick_column(cols, ("account", "gl account", "account name", "ledger")),
        "debit": _pick_column(cols, ("debit", "dr")),
        "credit": _pick_column(cols, ("credit", "cr")),
        "description": _pick_column(cols, ("description", "memo", "narration", "particular")),
    }
    if mapping["amount"] is None and mapping["debit"]:
        mapping["amount"] = mapping["debit"]
    return mapping


def _claude_map_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    if claude_client is None:
        return _detect_columns(df)
    sample = df.head(3).fillna("").astype(str).to_dict(orient="records")
    prompt = f"""Map journal entry columns to standard fields.
Columns: {list(df.columns)}
Sample rows: {json.dumps(sample)}

Return JSON only:
{{"date": "col or null", "voucher": "col or null", "party": "col or null", "trn": "col or null",
  "amount": "col or null", "vat_amount": "col or null", "type": "col or null",
  "account": "col or null", "debit": "col or null", "credit": "col or null", "description": "col or null"}}
Use exact column names from the file."""
    try:
        msg = claude_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        mapped = _json_from_claude(msg.content[0].text)
        out = _detect_columns(df)
        for key in out:
            val = mapped.get(key)
            if val and str(val).lower() in df.columns:
                out[key] = str(val).lower()
        return out
    except Exception as exc:
        logger.warning("Claude column mapping failed: %s", exc)
        return _detect_columns(df)


def _heuristic_category(text: str) -> str:
    t = (text or "").lower()
    if any(k in t for k in _BLOCKED_KWS) or any(k in t for k in _ENTERTAINMENT_KWS):
        return "NON_CLAIMABLE"
    if "reverse charge" in t or t.strip() in ("rc", "reverse_charge"):
        return "REVERSE_CHARGE"
    if "output" in t or "sales" in t or "collected" in t:
        return "OUTPUT_TAX"
    if "input" in t or "purchase" in t or "recoverable" in t or "deductible" in t:
        return "INPUT_TAX"
    if "vat payable" in t or "vat on sales" in t:
        return "OUTPUT_TAX"
    if "vat on purchase" in t:
        return "INPUT_TAX"
    return "NOT_VAT"


def _classify_accounts_with_claude(accounts: List[str]) -> Dict[str, str]:
    result = {a: _heuristic_category(a) for a in accounts}
    if not accounts or claude_client is None:
        return result
    prompt = f"""You are a UAE VAT accounting specialist. Classify each GL account name:
{json.dumps(accounts)}

Categories: INPUT_TAX, OUTPUT_TAX, REVERSE_CHARGE, NON_CLAIMABLE, NOT_VAT
Return JSON only: {{"Account Name": "CATEGORY"}}"""
    try:
        msg = claude_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        parsed = _json_from_claude(msg.content[0].text)
        for name, cat in parsed.items():
            cat_u = str(cat).upper().replace(" ", "_")
            if cat_u in ("INPUT_TAX", "OUTPUT_TAX", "REVERSE_CHARGE", "NON_CLAIMABLE", "NOT_VAT"):
                result[str(name)] = cat_u
        return result
    except Exception as exc:
        logger.warning("Claude account classification failed: %s", exc)
        return result


def _row_category(row: Dict[str, Any], account_map: Dict[str, str]) -> str:
    type_text = " ".join(
        str(row.get(k, "") or "")
        for k in ("type", "classification", "account", "description", "party")
    )
    cat = _heuristic_category(type_text)
    account = str(row.get("account") or "").strip()
    if account and account in account_map and account_map[account] != "NOT_VAT":
        return account_map[account]
    return cat


def _tax_claim_for(category: str, party: str) -> str:
    if category == "NON_CLAIMABLE":
        return "Blocked"
    if category == "REVERSE_CHARGE":
        return "Claimable"
    if category == "INPUT_TAX":
        if any(k in (party or "").lower() for k in _ENTERTAINMENT_KWS):
            return "Blocked"
        return "Claimable"
    if category == "OUTPUT_TAX":
        return "N/A"
    return "Non-Claimable"


def _display_classification(category: str) -> str:
    return {
        "INPUT_TAX": "Input Tax (Purchases)",
        "OUTPUT_TAX": "Output Tax (Sales)",
        "REVERSE_CHARGE": "Reverse Charge",
        "NON_CLAIMABLE": "Non-Claimable",
        "NOT_VAT": "Out of Scope",
    }.get(category, "Out of Scope")


def _parse_journal_rows(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for idx, raw in df.iterrows():
        if idx >= MAX_ROWS:
            break
        vat_amt = _to_float(raw.get(cols["vat_amount"])) if cols.get("vat_amount") else 0.0
        taxable = _to_float(raw.get(cols["amount"])) if cols.get("amount") else 0.0
        if cols.get("debit") or cols.get("credit"):
            taxable = max(_to_float(raw.get(cols.get("debit"))), _to_float(raw.get(cols.get("credit"))))
        if vat_amt == 0 and taxable == 0:
            continue
        party = str(raw.get(cols["party"], "") or "").strip() if cols.get("party") else ""
        trn = str(raw.get(cols["trn"], "") or "").strip() if cols.get("trn") else ""
        type_val = str(raw.get(cols["type"], "") or "").strip() if cols.get("type") else ""
        account = str(raw.get(cols["account"], "") or "").strip() if cols.get("account") else ""
        desc = str(raw.get(cols["description"], "") or "").strip() if cols.get("description") else ""
        date_val = raw.get(cols["date"]) if cols.get("date") else ""
        if isinstance(date_val, (pd.Timestamp, datetime, date)):
            date_str = date_val.isoformat()[:10] if hasattr(date_val, "isoformat") else str(date_val)
        else:
            date_str = str(date_val) if date_val and str(date_val) != "nan" else ""
        voucher = str(raw.get(cols["voucher"], "") or "").strip() if cols.get("voucher") else f"ROW-{idx + 1}"
        rate = round((vat_amt / taxable * 100), 2) if taxable > 0 and vat_amt > 0 else 5.0
        rows.append(
            {
                "date": date_str,
                "voucher_no": voucher,
                "party": party,
                "party_trn": trn,
                "taxable_amount": round(taxable, 2),
                "tax_rate_pct": rate,
                "vat_amount": round(vat_amt, 2),
                "type": type_val,
                "account": account,
                "description": desc,
                "debit": round(_to_float(raw.get(cols.get("debit"))), 2) if cols.get("debit") else 0.0,
                "credit": round(_to_float(raw.get(cols.get("credit"))), 2) if cols.get("credit") else 0.0,
            }
        )
    return rows


def _enrich_transactions(rows: List[Dict[str, Any]], account_map: Dict[str, str]) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for row in rows:
        category = _row_category(row, account_map)
        if category == "NOT_VAT" and row.get("vat_amount", 0) <= 0:
            continue
        party = row.get("party") or row.get("description") or ""
        claim = _tax_claim_for(category, party)
        flagged = (
            (row.get("party_trn") or "").upper() in ("", "NOT-REGISTERED", "N/A")
            or claim == "Blocked"
        )
        enriched.append(
            {
                **row,
                "category": category,
                "classification": _display_classification(category),
                "tax_claim": claim,
                "flagged": flagged,
            }
        )
    return enriched


def _sum_gl_vat(transactions: List[Dict[str, Any]]) -> Tuple[float, float]:
    gl_output = gl_input = 0.0
    for t in transactions:
        vat = float(t.get("vat_amount") or 0)
        cat = t.get("category")
        claim = t.get("tax_claim")
        if cat == "OUTPUT_TAX":
            gl_output += vat
        elif cat == "REVERSE_CHARGE":
            gl_output += vat
            gl_input += vat
        elif cat == "INPUT_TAX" and claim == "Claimable":
            gl_input += vat
    return round(gl_output, 2), round(gl_input, 2)


def _build_discrepancies(
    gl_input: float,
    gl_output: float,
    return_input: float,
    return_output: float,
    input_diff: float,
    output_diff: float,
    net_diff: float,
) -> List[Dict[str, Any]]:
    if abs(net_diff) <= RECON_TOLERANCE:
        return []
    items: List[Dict[str, Any]] = []
    if abs(output_diff) > RECON_TOLERANCE:
        items.append(
            {
                "title": "Output VAT Mismatch",
                "severity": "high",
                "gl_amount": gl_output,
                "return_amount": return_output,
                "difference": output_diff,
                "likely_cause": "Unposted sales invoice, advance payment not in GL, or output VAT coded to wrong account.",
                "action": "Review sales journal entries and VAT return Box 2.",
            }
        )
    if abs(input_diff) > RECON_TOLERANCE:
        items.append(
            {
                "title": "Input VAT Mismatch",
                "severity": "high",
                "gl_amount": gl_input,
                "return_amount": return_input,
                "difference": input_diff,
                "likely_cause": "Blocked entertainment VAT claimed, missing purchase invoices, or reverse charge not self-assessed.",
                "action": "Review purchase journals and VAT return Box 7.",
            }
        )
    if claude_client and items:
        try:
            prompt = f"""UAE VAT GL reconciliation discrepancy. GL input {gl_input}, return input {return_input}, GL output {gl_output}, return output {return_output}, net diff {net_diff}.
Suggest 2 likely causes and 1 action. Return JSON: {{"causes": ["..."], "action": "..."}}"""
            msg = claude_client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=512,
                messages=[{"role": "user", "content": prompt}],
            )
            ai = _json_from_claude(msg.content[0].text)
            for i, item in enumerate(items):
                if ai.get("causes"):
                    item["likely_cause"] = "; ".join(ai["causes"][:2])
                if ai.get("action"):
                    item["action"] = ai["action"]
        except Exception:
            pass
    return items


def _build_pdf(
    company_name: str,
    company_trn: str,
    period: str,
    movement: Dict[str, Any],
    transactions: List[Dict[str, Any]],
    discrepancies: List[Dict[str, Any]],
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    flow: List[Any] = []
    flow.append(Paragraph("UAE Tax — VAT Reconciliation vs Accounts", styles["Title"]))
    flow.append(Spacer(1, 3 * mm))
    flow.append(Paragraph(f"Company: {_pdf_text(company_name)}", styles["Normal"]))
    flow.append(Paragraph(f"TRN: {_pdf_text(company_trn or '—')}", styles["Normal"]))
    flow.append(Paragraph(f"Period: {_pdf_text(period)}", styles["Normal"]))
    flow.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}", styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))

    table_data = [
        ["VAT Movement Balance", "Amount (AED)"],
        ["Input Tax (Purchases) — from GL", _fmt_aed(movement["gl_input_vat"])],
        ["Output Tax (Sales) — from GL", _fmt_aed(movement["gl_output_vat"])],
        ["Net Tax Position (Current)", _fmt_aed(movement["gl_net_position"])],
        ["Input Tax (Purchases) — from Return", _fmt_aed(movement["return_input_vat"])],
        ["Output Tax (Sales) — from Return", _fmt_aed(movement["return_output_vat"])],
        ["Net Tax Position (Till Date)", _fmt_aed(movement["return_net"])],
        ["Input Difference", _fmt_aed(movement["input_difference"])],
        ["Output Difference", _fmt_aed(movement["output_difference"])],
        ["NET DIFFERENCE", _fmt_aed(movement["net_difference"])],
    ]
    t = Table(table_data, colWidths=[110 * mm, 50 * mm])
    hdr = colors.HexColor("#1E3A5F")
    net_color = colors.HexColor("#2DD4A0") if movement["is_reconciled"] else colors.HexColor("#F87171")
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), hdr),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, -1), (-1, -1), net_color),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ]
        )
    )
    flow.append(t)
    flow.append(Spacer(1, 4 * mm))

    if discrepancies:
        flow.append(Paragraph("<b>Discrepancies</b>", styles["Heading2"]))
        for d in discrepancies:
            flow.append(Paragraph(_pdf_text(f"{d.get('title')}: {_fmt_aed(d.get('difference', 0))}"), styles["Normal"]))
            flow.append(Paragraph(_pdf_text(d.get("likely_cause", "")), styles["Normal"]))
        flow.append(Spacer(1, 3 * mm))

    flow.append(Paragraph("<b>Recommendations</b>", styles["Heading2"]))
    if movement["is_reconciled"]:
        flow.append(Paragraph("GL VAT balances reconcile to the VAT return. Retain this report for FTA audit defence.", styles["Normal"]))
    else:
        flow.append(Paragraph("Investigate flagged transactions and repost journals before filing the VAT return.", styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))
    flow.append(Paragraph("Prepared by UAE Tax | For FTA audit defence use", styles["Normal"]))
    doc.build(flow)
    return buf.getvalue()


def _build_excel(movement: Dict[str, Any], transactions: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)

    ws1 = wb.active
    ws1.title = "VAT Movement"
    ws1.append(["Line", "Amount (AED)"])
    for cell in ws1[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for label, key in [
        ("GL Input VAT", "gl_input_vat"),
        ("GL Output VAT", "gl_output_vat"),
        ("GL Net", "gl_net_position"),
        ("Return Input VAT", "return_input_vat"),
        ("Return Output VAT", "return_output_vat"),
        ("Return Net", "return_net"),
        ("Input Difference", "input_difference"),
        ("Output Difference", "output_difference"),
        ("Net Difference", "net_difference"),
    ]:
        ws1.append([label, movement.get(key, 0)])

    ws2 = wb.create_sheet("Transactions")
    headers = [
        "Date", "Voucher No", "Party", "Party TRN", "Taxable Amount",
        "Tax Rate %", "VAT Amount", "Classification", "Tax Claim", "DR", "CR", "Flagged",
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for t in transactions:
        ws2.append(
            [
                t.get("date"),
                t.get("voucher_no"),
                t.get("party"),
                t.get("party_trn"),
                t.get("taxable_amount"),
                t.get("tax_rate_pct"),
                t.get("vat_amount"),
                t.get("classification"),
                t.get("tax_claim"),
                t.get("debit"),
                t.get("credit"),
                t.get("flagged"),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _save_reconciliation(
    db: Session,
    company_id: int,
    period: str,
    company_trn: str,
    movement: Dict[str, Any],
    transactions: List[Dict[str, Any]],
    discrepancies: List[Dict[str, Any]],
) -> str:
    recon_id = str(uuid.uuid4())
    row = VATAccountsReconciliation(
        id=recon_id,
        company_id=company_id,
        period=period,
        company_trn=company_trn,
        gl_input_vat=movement["gl_input_vat"],
        gl_output_vat=movement["gl_output_vat"],
        return_input_vat=movement["return_input_vat"],
        return_output_vat=movement["return_output_vat"],
        net_difference=movement["net_difference"],
        is_reconciled=movement["is_reconciled"],
        transactions=transactions,
        discrepancies=discrepancies,
    )
    db.add(row)
    db.commit()

    sb = _get_supabase()
    if sb:
        try:
            sb.table("vat_accounts_reconciliations").insert(
                {
                    "id": recon_id,
                    "company_id": company_id,
                    "period": period,
                    "company_trn": company_trn,
                    "gl_input_vat": movement["gl_input_vat"],
                    "gl_output_vat": movement["gl_output_vat"],
                    "return_input_vat": movement["return_input_vat"],
                    "return_output_vat": movement["return_output_vat"],
                    "net_difference": movement["net_difference"],
                    "is_reconciled": movement["is_reconciled"],
                    "transactions": transactions,
                    "discrepancies": discrepancies,
                }
            ).execute()
        except Exception as exc:
            logger.warning("Supabase vat_accounts_reconciliations insert failed: %s", exc)
    return recon_id


@router.get("/returns")
async def list_returns_for_recon(
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
    limit: int = Query(20, ge=1, le=100),
):
    """VAT returns with Box 2/7/8 for reconciliation dropdown."""
    rows = (
        db.query(VATReturn)
        .filter(VATReturn.company_id == company_id)
        .order_by(VATReturn.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "return_id": r.id,
            "period_start": r.period_start.isoformat() if r.period_start else None,
            "period_end": r.period_end.isoformat() if r.period_end else None,
            "output_vat": float(r.box2_vat_on_supplies or 0),
            "input_vat": float(r.box7_vat_on_expenses or 0),
            "net_vat": float(r.box8_vat_payable_or_refundable or 0),
            "status": r.status,
        }
        for r in rows
    ]


@router.get("/history")
async def reconciliation_history(
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
    limit: int = Query(10, ge=1, le=50),
):
    rows = (
        db.query(VATAccountsReconciliation)
        .filter(VATAccountsReconciliation.company_id == company_id)
        .order_by(VATAccountsReconciliation.created_at.desc())
        .limit(limit)
        .all()
    )
    return {
        "reconciliations": [
            {
                "id": r.id,
                "period": r.period,
                "net_difference": r.net_difference,
                "is_reconciled": r.is_reconciled,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]
    }


@router.post("/reconcile")
async def reconcile_vat_accounts(
    file: UploadFile = File(...),
    period: str = Form(...),
    company_trn: str = Form(""),
    vat_return_id: Optional[int] = Form(None),
    manual_output_vat: Optional[float] = Form(None),
    manual_input_vat: Optional[float] = Form(None),
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(raw) > MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds 10MB limit")

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    lower = file.filename.lower()
    buf = BytesIO(raw)
    try:
        if lower.endswith(".csv"):
            df = pd.read_csv(buf)
        elif lower.endswith((".xlsx", ".xls")):
            df = pd.read_excel(buf, engine="openpyxl")
        else:
            raise HTTPException(status_code=400, detail="Unsupported format. Use .csv, .xlsx, or .xls")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read file: {exc}") from exc

    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no data rows")
    if len(df) > MAX_ROWS:
        df = df.head(MAX_ROWS)

    cols = _detect_columns(df)
    if not cols.get("vat_amount") and not cols.get("amount"):
        cols = _claude_map_columns(df)
    if not cols.get("vat_amount") and not cols.get("amount"):
        raise HTTPException(status_code=400, detail=f"Could not detect VAT or amount columns. Found: {list(df.columns)}")

    journal_rows = _parse_journal_rows(df, cols)
    if not journal_rows:
        raise HTTPException(status_code=400, detail="No VAT journal entries found in file")

    accounts = sorted({r["account"] for r in journal_rows if r.get("account")})
    account_map = _classify_accounts_with_claude(accounts)
    transactions = _enrich_transactions(journal_rows, account_map)

    gl_output, gl_input = _sum_gl_vat(transactions)
    gl_net = round(gl_output - gl_input, 2)

    return_output = return_input = return_net = 0.0
    if vat_return_id:
        vat_return = (
            db.query(VATReturn)
            .filter(VATReturn.id == vat_return_id, VATReturn.company_id == company_id)
            .first()
        )
        if not vat_return:
            raise HTTPException(status_code=404, detail="VAT return not found")
        return_output = float(vat_return.box2_vat_on_supplies or 0)
        return_input = float(vat_return.box7_vat_on_expenses or 0)
        return_net = float(vat_return.box8_vat_payable_or_refundable or 0)
    else:
        if manual_output_vat is None or manual_input_vat is None:
            raise HTTPException(
                status_code=400,
                detail="Select a VAT return or enter manual Output VAT and Input VAT amounts",
            )
        return_output = float(manual_output_vat)
        return_input = float(manual_input_vat)
        return_net = round(return_output - return_input, 2)

    input_difference = round(gl_input - return_input, 2)
    output_difference = round(gl_output - return_output, 2)
    net_difference = round(gl_net - return_net, 2)
    is_reconciled = abs(net_difference) <= RECON_TOLERANCE

    movement = {
        "gl_input_vat": gl_input,
        "gl_output_vat": gl_output,
        "gl_net_position": gl_net,
        "return_input_vat": round(return_input, 2),
        "return_output_vat": round(return_output, 2),
        "return_net": round(return_net, 2),
        "input_difference": input_difference,
        "output_difference": output_difference,
        "net_difference": net_difference,
        "is_reconciled": is_reconciled,
    }

    discrepancies = _build_discrepancies(
        gl_input, gl_output, return_input, return_output,
        input_difference, output_difference, net_difference,
    )

    pdf_bytes = b""
    excel_bytes = b""
    try:
        pdf_bytes = _build_pdf(
            company.name, company_trn or company.trn or "", period,
            movement, transactions, discrepancies,
        )
    except Exception as exc:
        logger.warning("PDF build failed: %s", exc)
    try:
        excel_bytes = _build_excel(movement, transactions)
    except Exception as exc:
        logger.warning("Excel build failed: %s", exc)

    recon_id = _save_reconciliation(
        db, company_id, period, company_trn or company.trn or "",
        movement, transactions, discrepancies,
    )

    return {
        "vat_movement": movement,
        "transactions": transactions,
        "discrepancies": discrepancies,
        "row_count": len(transactions),
        "pdf_base64": base64.b64encode(pdf_bytes).decode("ascii") if pdf_bytes else "",
        "excel_base64": base64.b64encode(excel_bytes).decode("ascii") if excel_bytes else "",
        "reconciliation_id": recon_id,
    }


@router.patch("/{reconciliation_id}/mark-reconciled")
async def mark_reconciled(
    reconciliation_id: str,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    row = (
        db.query(VATAccountsReconciliation)
        .filter(
            VATAccountsReconciliation.id == reconciliation_id,
            VATAccountsReconciliation.company_id == company_id,
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Reconciliation not found")
    row.is_reconciled = True
    row.net_difference = 0.0
    db.commit()
    return {"ok": True, "reconciliation_id": reconciliation_id}
