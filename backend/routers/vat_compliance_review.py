"""VAT Compliance Review — full FTA compliance analysis with Claude, PDF/Excel export."""
from __future__ import annotations

import base64
import json
import logging
import os
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional
from xml.sax.saxutils import escape

import pandas as pd
from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from database import get_db
from middleware.auth import get_current_company_id, require_role
from models import Company, Transaction, VATComplianceReview
from utils.audit import log_ai_audit
from utils.prompt_guard import sanitize_input

load_dotenv()

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/vat-compliance", tags=["VAT Compliance Review"])

MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 500

anthropic_api_key = os.getenv("ANTHROPIC_API_KEY")
claude_client = Anthropic(api_key=anthropic_api_key) if anthropic_api_key else None

COMPLIANCE_SYSTEM_PROMPT = """You are a UAE VAT Compliance Specialist with expertise in Federal Tax Authority (FTA) regulations and VAT201 return filing. Analyse the transaction data provided and return a JSON compliance report.

UAE VAT RULES TO APPLY:
- Standard rate: 5% on taxable supplies
- Zero rate: exports, international transport, healthcare, education, residential (first supply)
- Exempt: bare land, local passenger transport, financial services (margin-based)
- Out of scope: salaries, dividends
- Reverse charge: imported services
- Blocked input tax: entertainment, personal-use motor vehicles, employee personal benefits
- Valid tax invoice requires: supplier name + TRN, customer TRN (B2B > AED 10,000), invoice date + number, description, taxable amount, VAT rate, VAT amount, total with VAT
- Simplified invoice: only for supplies < AED 10,000

Analyse ALL transactions provided and return ONLY valid JSON in this exact structure:

{
  "executive_summary": {
    "period": "",
    "total_output_vat": 0.0,
    "total_input_vat": 0.0,
    "net_vat_position": 0.0,
    "net_vat_label": "PAYABLE" or "REFUNDABLE",
    "issues_count": 0,
    "compliance_rating": "COMPLIANT" or "MINOR ISSUES" or "REQUIRES CORRECTION" or "HIGH RISK",
    "cfo_summary": ""
  },
  "findings": [
    {
      "number": 1,
      "area": "",
      "issue": "",
      "transaction_ref": "",
      "amount_aed": 0.0,
      "vat_impact_aed": 0.0,
      "risk": "HIGH" or "MEDIUM" or "LOW",
      "action": ""
    }
  ],
  "vat201": {
    "box_1a_standard_rated": 0.0,
    "box_1b_zero_rated": 0.0,
    "box_2_exempt": 0.0,
    "box_3_output_vat": 0.0,
    "box_9_input_vat": 0.0,
    "box_10_net_vat": 0.0
  },
  "priority_actions": [
    "Action 1 description"
  ],
  "audit_triggers": {
    "advance_payments": false,
    "intercompany": false,
    "deemed_supplies": false,
    "bad_debt": false,
    "late_registration_risk": false,
    "deregistration_risk": false,
    "blocked_input_tax": false,
    "missing_trn": false
  },
  "high_risk_found": false,
  "disclosure_needed": false
}

Set audit_triggers fields to TRUE when that trigger is flagged as an issue/risk, FALSE when clean.
Reduce box_9_input_vat for blocked input tax that should not be recovered."""


class DraftDisclosureRequest(BaseModel):
    review_id: str
    company_name: str
    company_trn: str
    findings_summary: str
    period: str


def _get_supabase():
    url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
    key = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_KEY") or "").strip()
    if not url or not key:
        return None
    try:
        from supabase import create_client  # type: ignore

        return create_client(url, key)
    except Exception as exc:
        logger.warning("Supabase client init failed: %s", exc)
        return None


def _json_from_claude_text(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    if "```json" in cleaned:
        cleaned = cleaned.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in cleaned:
        cleaned = cleaned.split("```", 1)[1].split("```", 1)[0].strip()
    return json.loads(cleaned)


def _pdf_text(value: Any) -> str:
    """Escape dynamic text for ReportLab Paragraph (XML-safe)."""
    return escape(str(value or ""))


def _excel_cell(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value.isoformat() if hasattr(value, "isoformat") else str(value)
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def _normalize_analysis(analysis: Dict[str, Any]) -> Dict[str, Any]:
    """Ensure required keys exist so PDF/UI never crash on partial Claude output."""
    merged = dict(analysis or {})
    es = dict(merged.get("executive_summary") or {})
    es.setdefault("period", "")
    es.setdefault("total_output_vat", 0)
    es.setdefault("total_input_vat", 0)
    es.setdefault("net_vat_position", 0)
    es.setdefault("net_vat_label", "payable")
    es.setdefault("issues_count", len(merged.get("findings") or []))
    es.setdefault("compliance_rating", "REQUIRES CORRECTION")
    es.setdefault("cfo_summary", "")
    merged["executive_summary"] = es
    merged.setdefault("findings", [])
    merged.setdefault("vat201", {})
    merged.setdefault("priority_actions", [])
    merged.setdefault("audit_triggers", {})
    merged.setdefault("high_risk_found", es.get("compliance_rating") == "HIGH RISK")
    merged.setdefault("disclosure_needed", bool(merged.get("high_risk_found")))
    return merged


_HEADER_KEYWORDS = {
    "desc", "amount", "date", "vendor", "supplier", "customer", "invoice",
    "transaction", "type", "vat", "ref", "narration", "particular", "net", "total", "rate",
    "treatment", "trn",
}


def _header_row_match_count(columns: list) -> int:
    cleaned = [str(c).strip().lower() for c in columns if str(c).strip().lower() not in ("", "nan")]
    if not cleaned:
        return 0
    return sum(1 for kw in _HEADER_KEYWORDS if any(kw in cell for cell in cleaned))


def _is_valid_excel_header(columns: list) -> bool:
    cleaned = [str(c).strip().lower() for c in columns if str(c).strip().lower() not in ("", "nan")]
    if not cleaned:
        return False
    unnamed = sum(1 for c in cleaned if c.startswith("unnamed"))
    if unnamed >= max(1, len(cleaned) // 2):
        return False
    return _header_row_match_count(columns) >= 3


def _looks_like_title_row(row_values) -> bool:
    vals = [str(v).strip() for v in row_values if pd.notna(v) and str(v).strip().lower() not in ("", "nan")]
    if not vals:
        return False
    text = " ".join(vals).lower()
    # Single merged title row (one cell filled) or subtitle with company/period text
    if len(vals) <= 2 and any(
        hint in text
        for hint in ("llc", "ltd", "inc", " fz", " q1", " q2", " q3", " q4", "transactions", "—", "–", "report")
    ):
        return True
    # Row has title-like text but lacks real column header names
    has_header_names = any(k in text for k in ("date", "description", "amount", "vendor", "invoice no", "invoice ref"))
    if not has_header_names and any(hint in text for hint in ("llc", "ltd", "real estate", "trading", "contracting")):
        return True
    return False


def _read_excel_with_header_detection(buf: BytesIO) -> pd.DataFrame:
    raw_df = pd.read_excel(buf, engine="openpyxl", header=None)
    if raw_df.empty:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_row: Optional[int] = None
    for i in range(min(8, len(raw_df))):
        row_vals = raw_df.iloc[i].values
        if _looks_like_title_row(row_vals):
            continue
        candidate_cols = [str(v).strip() for v in row_vals]
        if _is_valid_excel_header(candidate_cols):
            header_row = i
            break

    if header_row is None:
        for i in range(min(3, len(raw_df))):
            candidate_cols = [str(v).strip() for v in raw_df.iloc[i].values]
            if _is_valid_excel_header(candidate_cols):
                header_row = i
                break

    if header_row is None:
        raise HTTPException(
            status_code=400,
            detail="Could not detect column headers — ensure row 1–3 contains Date, Description, Amount, Vendor, or VAT columns",
        )

    df = raw_df.iloc[header_row + 1 :].copy()
    df.columns = [str(v).strip() for v in raw_df.iloc[header_row].values]
    return df.reset_index(drop=True)


def _read_upload(file: UploadFile) -> pd.DataFrame:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    raw = file.file.read()
    if len(raw) > MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds 10MB limit")
    lower = file.filename.lower()
    buf = BytesIO(raw)
    if lower.endswith(".csv"):
        df = pd.read_csv(buf)
    elif lower.endswith((".xlsx", ".xls")):
        df = _read_excel_with_header_detection(buf)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Upload .xlsx, .xls, or .csv")
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    if len(df) > MAX_ROWS:
        df = df.head(MAX_ROWS)
    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no transaction rows")
    return df


def _detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    mapping: Dict[str, Optional[str]] = {
        "date": None,
        "description": None,
        "amount": None,
        "vat_amount": None,
        "transaction_type": None,
        "vendor": None,
        "invoice_ref": None,
        "vat_rate": None,
    }
    for col in df.columns:
        cl = str(col).lower()
        if mapping["date"] is None and ("date" in cl or cl == "invoice date"):
            mapping["date"] = col
        if mapping["description"] is None and any(k in cl for k in ("desc", "narration", "particular")):
            mapping["description"] = col
        if mapping["amount"] is None and (
            cl in ("amount aed", "amount_aed", "net amount", "net_amount", "taxable amount")
            or ("amount" in cl and "vat" not in cl and "total" not in cl)
        ):
            mapping["amount"] = col
        if mapping["vat_amount"] is None and any(k in cl for k in ("vat amount", "vat_amount", "tax amount", "tax_amount")):
            mapping["vat_amount"] = col
        if mapping["transaction_type"] is None and (
            ("transaction" in cl and "type" in cl) or cl in ("type", "treatment", "vat treatment", "vat_treatment")
        ):
            mapping["transaction_type"] = col
        if mapping["vendor"] is None and any(k in cl for k in ("vendor", "supplier", "customer")):
            mapping["vendor"] = col
        if mapping["invoice_ref"] is None and any(k in cl for k in ("invoice", "ref", "reference", "doc no")):
            mapping["invoice_ref"] = col
        if mapping["vat_rate"] is None and ("vat rate" in cl or cl == "rate"):
            mapping["vat_rate"] = col
    return mapping


def _claude_map_columns(df: pd.DataFrame, db: Optional[Session] = None, company_id: Optional[int] = None) -> Dict[str, Optional[str]]:
    if claude_client is None:
        raise HTTPException(status_code=503, detail="Claude API not configured for column mapping")
    sample = df.head(3).fillna("").astype(str).to_dict(orient="records")
    msg = claude_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": (
                    "Map these spreadsheet columns to transaction fields. "
                    "Return ONLY JSON with keys: date, description, amount, vat_amount, "
                    "transaction_type, vendor, invoice_ref, vat_rate. "
                    f"Values must be exact column names from: {list(df.columns)}. "
                    f"Sample rows: {json.dumps(sample)}"
                ),
            }
        ],
    )
    if db is not None:
        try:
            log_ai_audit(
                db,
                company_id=company_id,
                user_email="user",
                action_type="ai_call",
                feature="vat_compliance_review",
                input_summary=f"Column mapping for {len(df.columns)} columns",
                output_summary=str(msg.content[0].text)[:100],
                status="success",
            )
        except Exception:
            pass
    return _json_from_claude_text(msg.content[0].text)


def _to_float(val: Any) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("AED", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _normalize_transactions(df: pd.DataFrame, cols: Dict[str, Optional[str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for idx, row in df.iterrows():
        desc_col = cols.get("description")
        if not desc_col:
            continue
        description = str(row.get(desc_col, "")).strip()
        if not description:
            continue
        tx_type = "purchase"
        if cols.get("transaction_type"):
            raw_type = str(row.get(cols["transaction_type"], "")).lower()
            if "sale" in raw_type:
                tx_type = "sale"
        ref = ""
        if cols.get("invoice_ref") and pd.notna(row.get(cols["invoice_ref"])):
            ref = str(row.get(cols["invoice_ref"])).strip()
        elif cols.get("date"):
            ref = f"ROW-{int(idx) + 1}"
        rows.append(
            {
                "row": int(idx) + 1,
                "date": str(row.get(cols["date"], "")) if cols.get("date") else "",
                "description": description,
                "vendor_or_customer": str(row.get(cols["vendor"], "")) if cols.get("vendor") else "",
                "amount_aed": _to_float(row.get(cols["amount"])) if cols.get("amount") else 0.0,
                "vat_amount_aed": _to_float(row.get(cols["vat_amount"])) if cols.get("vat_amount") else 0.0,
                "vat_rate": str(row.get(cols["vat_rate"], "")) if cols.get("vat_rate") else "",
                "transaction_type": tx_type,
                "invoice_ref": ref,
            }
        )
    return rows


def transactions_for_compliance_from_db(db: Session, company_id: int) -> List[Dict[str, Any]]:
    """Build compliance analyse payload from all stored transactions for a company."""
    txns = (
        db.query(Transaction)
        .filter(Transaction.company_id == company_id)
        .order_by(Transaction.date, Transaction.id)
        .all()
    )
    rows: List[Dict[str, Any]] = []
    for i, t in enumerate(txns, start=1):
        rows.append(
            {
                "row": i,
                "date": t.date.isoformat() if t.date else "",
                "description": t.description or "",
                "vendor_or_customer": t.vendor_or_customer or "",
                "amount_aed": float(t.amount_aed or 0),
                "vat_amount_aed": float(t.vat_amount_aed or 0),
                "vat_rate": "5%",
                "transaction_type": (t.transaction_type or "purchase").lower(),
                "invoice_ref": t.invoice_number or f"ROW-{i}",
            }
        )
    return rows


def _run_compliance_analysis(
    transactions: List[Dict[str, Any]],
    period: str,
    company_trn: str,
    entity_type: str,
    db: Optional[Session] = None,
    company_id: Optional[int] = None,
    actor: str = "user",
) -> Dict[str, Any]:
    if claude_client is None:
        raise HTTPException(status_code=503, detail="Claude API not configured")
    period = sanitize_input(period or "", "period")
    company_trn = sanitize_input(company_trn or "", "company_trn")
    system = COMPLIANCE_SYSTEM_PROMPT.replace("{entity_type}", entity_type).replace("{period}", period).replace(
        "{company_trn}", company_trn
    )
    user_msg = f"ENTITY TYPE: {entity_type}\nPERIOD: {period}\nCOMPANY TRN: {company_trn}\n\nAnalyse these {len(transactions)} transactions for UAE VAT compliance:\n{json.dumps(transactions, default=str)}"
    try:
        msg = claude_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=8192,
            system=system,
            messages=[{"role": "user", "content": user_msg}],
        )
        analysis = _normalize_analysis(_json_from_claude_text(msg.content[0].text))
        if db is not None:
            try:
                risk = analysis.get("executive_summary", {}).get("overall_risk_rating", "N/A")
                log_ai_audit(
                    db,
                    company_id=company_id,
                    user_email=actor,
                    action_type="ai_call",
                    feature="vat_compliance_review",
                    input_summary=f"Analysed {len(transactions)} transactions for {period}",
                    output_summary=f"Risk: {risk}, issues: {analysis.get('executive_summary', {}).get('issues_count', 'N/A')}",
                    status="success",
                )
            except Exception:
                pass
        return analysis
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="Claude returned invalid JSON for compliance analysis") from exc
    except Exception as exc:
        logger.exception("Claude compliance analysis failed")
        raise HTTPException(status_code=500, detail=f"Compliance analysis failed: {exc}") from exc


def _fmt_aed(amount: float) -> str:
    return f"AED {amount:,.2f}"


def _build_pdf(
    analysis: Dict[str, Any],
    company_name: str,
    company_trn: str,
    period: str,
    entity_type: str,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    es = analysis.get("executive_summary", {})
    flow: List[Any] = []
    flow.append(Paragraph("UAE Tax — UAE VAT Compliance Review", styles["Title"]))
    flow.append(Spacer(1, 3 * mm))
    flow.append(Paragraph(f"Company: {_pdf_text(company_name)}", styles["Normal"]))
    flow.append(Paragraph(f"TRN: {_pdf_text(company_trn or '—')}", styles["Normal"]))
    flow.append(Paragraph(f"Period: {_pdf_text(period)}", styles["Normal"]))
    flow.append(Paragraph(f"Entity: {_pdf_text(entity_type)}", styles["Normal"]))
    flow.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}", styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))

    flow.append(Paragraph("<b>Executive Summary</b>", styles["Heading2"]))
    flow.append(Paragraph(f"Rating: {_pdf_text(es.get('compliance_rating', '—'))}", styles["Normal"]))
    flow.append(Paragraph(f"Output VAT: {_fmt_aed(float(es.get('total_output_vat', 0)))}", styles["Normal"]))
    flow.append(Paragraph(f"Input VAT: {_fmt_aed(float(es.get('total_input_vat', 0)))}", styles["Normal"]))
    flow.append(Paragraph(f"Net VAT ({_pdf_text(es.get('net_vat_label', ''))}): {_fmt_aed(float(es.get('net_vat_position', 0)))}", styles["Normal"]))
    flow.append(Paragraph(f"Issues: {es.get('issues_count', 0)}", styles["Normal"]))
    if es.get("cfo_summary"):
        flow.append(Spacer(1, 2 * mm))
        flow.append(Paragraph(_pdf_text(es["cfo_summary"]), styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))

    findings = analysis.get("findings", [])
    flow.append(Paragraph("<b>Compliance Findings</b>", styles["Heading2"]))
    if findings:
        fdata = [["#", "Area", "Issue", "Ref", "VAT Impact", "Risk"]]
        for f in findings:
            fdata.append(
                [
                    str(f.get("number", "")),
                    str(f.get("area", ""))[:30],
                    str(f.get("issue", ""))[:40],
                    str(f.get("transaction_ref", ""))[:15],
                    _fmt_aed(float(f.get("vat_impact_aed", 0))),
                    str(f.get("risk", "")),
                ]
            )
        ft = Table(fdata, colWidths=[8 * mm, 28 * mm, 45 * mm, 22 * mm, 28 * mm, 18 * mm])
        ft.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.3, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white)]))
        flow.append(ft)
    else:
        flow.append(Paragraph("No issues found.", styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))

    vat201 = analysis.get("vat201", {})
    flow.append(Paragraph("<b>VAT201 Return Summary</b>", styles["Heading2"]))
    vdata = [
        ["Box", "Description", "Amount (AED)"],
        ["1a", "Standard rated supplies", _fmt_aed(float(vat201.get("box_1a_standard_rated", 0)))],
        ["1b", "Zero-rated supplies", _fmt_aed(float(vat201.get("box_1b_zero_rated", 0)))],
        ["2", "Exempt supplies", _fmt_aed(float(vat201.get("box_2_exempt", 0)))],
        ["3", "Output VAT", _fmt_aed(float(vat201.get("box_3_output_vat", 0)))],
        ["9", "Recoverable input VAT", _fmt_aed(float(vat201.get("box_9_input_vat", 0)))],
        ["10", "NET VAT PAYABLE / (REFUNDABLE)", _fmt_aed(float(vat201.get("box_10_net_vat", 0)))],
    ]
    vt = Table(vdata, colWidths=[15 * mm, 80 * mm, 50 * mm])
    vt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.3, colors.grey), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C8A951"))]))
    flow.append(vt)
    flow.append(Spacer(1, 4 * mm))

    actions = analysis.get("priority_actions", [])
    flow.append(Paragraph("<b>Priority Actions Before Filing</b>", styles["Heading2"]))
    for i, act in enumerate(actions, 1):
        flow.append(Paragraph(f"{i}. {_pdf_text(act)}", styles["Normal"]))
    flow.append(Spacer(1, 4 * mm))

    triggers = analysis.get("audit_triggers", {})
    flow.append(Paragraph("<b>FTA Audit Triggers — Status</b>", styles["Heading2"]))
    labels = {
        "advance_payments": "VAT on advance payments received",
        "intercompany": "VAT on intercompany transactions",
        "deemed_supplies": "Deemed supplies (gifts/samples > AED 500)",
        "bad_debt": "Bad debt adjustment claims",
        "late_registration_risk": "Late registration penalty risk",
        "deregistration_risk": "De-registration threshold check",
        "blocked_input_tax": "Blocked input tax (vehicles, entertainment)",
        "missing_trn": "Missing TRN on supplier invoices",
    }
    for key, label in labels.items():
        flagged = bool(triggers.get(key, False))
        status = "FLAGGED" if flagged else "CLEAN"
        flow.append(Paragraph(f"• {label}: {status}", styles["Normal"]))

    flow.append(Spacer(1, 6 * mm))
    flow.append(Paragraph("Prepared by UAE Tax | For review by qualified UAE VAT professional before submission", styles["Normal"]))
    doc.build(flow)
    return buf.getvalue()


def _build_excel(analysis: Dict[str, Any], transactions: List[Dict[str, Any]], df: pd.DataFrame) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    vat201 = analysis.get("vat201", {})

    ws1 = wb.active
    ws1.title = "VAT201 Return"
    ws1.append(["Box", "Description", "Amount (AED)"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    rows = [
        ("1a", "Standard rated supplies", float(vat201.get("box_1a_standard_rated", 0))),
        ("1b", "Zero-rated supplies", float(vat201.get("box_1b_zero_rated", 0))),
        ("2", "Exempt supplies", float(vat201.get("box_2_exempt", 0))),
        ("3", "Output VAT", float(vat201.get("box_3_output_vat", 0))),
        ("9", "Recoverable input VAT", float(vat201.get("box_9_input_vat", 0))),
        ("10", "NET VAT PAYABLE / (REFUNDABLE)", float(vat201.get("box_10_net_vat", 0))),
    ]
    for r in rows:
        ws1.append(list(r))

    ws2 = wb.create_sheet("Compliance Findings")
    ws2.append(["#", "Area", "Issue Found", "Transaction Ref", "Amount (AED)", "VAT Impact (AED)", "Risk", "Action Required"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for f in analysis.get("findings", []):
        ws2.append(
            [
                f.get("number"),
                f.get("area"),
                f.get("issue"),
                f.get("transaction_ref"),
                f.get("amount_aed"),
                f.get("vat_impact_aed"),
                f.get("risk"),
                f.get("action"),
            ]
        )

    ws3 = wb.create_sheet("Transactions")
    out_df = df.copy()
    out_df["ai_vat_treatment"] = ""
    ref_to_treatment: Dict[str, str] = {}
    for f in analysis.get("findings", []):
        ref = str(f.get("transaction_ref", ""))
        if ref:
            ref_to_treatment[ref] = str(f.get("area", "review"))
    if "invoice_ref" in [c.lower() for c in out_df.columns]:
        ref_col = next(c for c in out_df.columns if "invoice" in c.lower() or "ref" in c.lower())
        for i, val in enumerate(out_df[ref_col].astype(str)):
            out_df.at[out_df.index[i], "ai_vat_treatment"] = ref_to_treatment.get(val, "standard")
    else:
        invoice_col = next((c for c in out_df.columns if "invoice" in c.lower()), None)
        if invoice_col:
            for i, val in enumerate(out_df[invoice_col].astype(str)):
                out_df.at[out_df.index[i], "ai_vat_treatment"] = ref_to_treatment.get(val, "standard")
        else:
            for i in range(len(out_df)):
                out_df.at[out_df.index[i], "ai_vat_treatment"] = "review"
    for r in out_df.itertuples(index=False):
        ws3.append([_excel_cell(v) for v in r])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _save_review_db(
    db: Session,
    company_id: int,
    period: str,
    company_trn: str,
    entity_type: str,
    analysis: Dict[str, Any],
    row_count: int,
    review_id: str,
    source: str = "manual",
) -> str:
    es = analysis.get("executive_summary", {})
    row = VATComplianceReview(
        id=review_id,
        company_id=company_id,
        period=period,
        company_trn=company_trn,
        entity_type=entity_type,
        compliance_rating=es.get("compliance_rating"),
        issues_count=int(es.get("issues_count", 0) or 0),
        net_vat_position=float(es.get("net_vat_position", 0) or 0),
        output_vat=float(es.get("total_output_vat", 0) or 0),
        input_vat=float(es.get("total_input_vat", 0) or 0),
        analysis=analysis,
        row_count=row_count,
        source=source,
    )
    db.add(row)
    db.commit()
    return review_id


def _save_review_supabase(
    company_id: int,
    period: str,
    company_trn: str,
    entity_type: str,
    analysis: Dict[str, Any],
    db: Optional[Session] = None,
    row_count: int = 0,
    source: str = "manual",
) -> str:
    review_id = str(uuid.uuid4())
    if db is not None:
        try:
            _save_review_db(
                db, company_id, period, company_trn, entity_type,
                analysis, row_count, review_id, source=source,
            )
        except Exception as exc:
            logger.warning("SQLite vat_compliance_reviews insert failed: %s", exc)
            db.rollback()
    sb = _get_supabase()
    if not sb:
        return review_id
    es = analysis.get("executive_summary", {})
    try:
        sb.table("vat_compliance_reviews").insert(
            {
                "id": review_id,
                "company_id": company_id,
                "period": period,
                "company_trn": company_trn,
                "entity_type": entity_type,
                "compliance_rating": es.get("compliance_rating"),
                "issues_count": es.get("issues_count", 0),
                "net_vat_position": es.get("net_vat_position", 0),
                "output_vat": es.get("total_output_vat", 0),
                "input_vat": es.get("total_input_vat", 0),
                "findings": analysis.get("findings", []),
                "vat201": analysis.get("vat201", {}),
                "audit_triggers": analysis.get("audit_triggers", {}),
            }
        ).execute()
    except Exception as exc:
        logger.warning("Supabase vat_compliance_reviews insert failed: %s", exc)
    return review_id


def _export_compliance_reports(
    analysis: Dict[str, Any],
    company_name: str,
    company_trn: str,
    period: str,
    entity_type: str,
    transactions: List[Dict[str, Any]],
    df: Optional[pd.DataFrame] = None,
) -> tuple[str, str]:
    pdf_b64 = ""
    excel_b64 = ""
    try:
        pdf_b64 = base64.b64encode(
            _build_pdf(analysis, company_name, company_trn, period, entity_type)
        ).decode("ascii")
    except Exception as exc:
        logger.warning("Compliance PDF export failed: %s", exc)
    try:
        excel_b64 = base64.b64encode(
            _build_excel(analysis, transactions, df if df is not None else pd.DataFrame())
        ).decode("ascii")
    except Exception as exc:
        logger.warning("Compliance Excel export failed: %s", exc)
    return pdf_b64, excel_b64


def _run_and_persist_compliance(
    *,
    transactions: List[Dict[str, Any]],
    period: str,
    company_trn: str,
    entity_type: str,
    company_id: int,
    company_name: str,
    db: Session,
    actor: str,
    source: str,
    df: Optional[pd.DataFrame] = None,
    include_exports: bool = True,
) -> Dict[str, Any]:
    logger.info("Running Claude compliance analysis on %s transactions", len(transactions))
    analysis = _run_compliance_analysis(
        transactions, period, company_trn, entity_type, db, company_id, actor
    )
    review_id = _save_review_supabase(
        company_id, period, company_trn, entity_type, analysis,
        db=db, row_count=len(transactions), source=source,
    )
    pdf_b64 = excel_b64 = ""
    if include_exports:
        pdf_b64, excel_b64 = _export_compliance_reports(
            analysis, company_name, company_trn, period, entity_type, transactions, df
        )
    return {
        "analysis": analysis,
        "row_count": len(transactions),
        "pdf_base64": pdf_b64,
        "excel_base64": excel_b64,
        "review_id": review_id,
    }


@router.post("/analyse")
async def analyse_compliance(
    file: UploadFile = File(...),
    period: str = Form(...),
    company_trn: str = Form(""),
    entity_type: str = Form("Mainland UAE"),
    auth: dict = require_role("analyst"),
    db: Session = Depends(get_db),
):
    company_id = auth["company_id"]
    actor = auth["user"].get("email") or "user"
    logger.info("VAT compliance analyse started company_id=%s file=%s", company_id, file.filename)
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    raw = await file.read()
    if not file.filename:
        raise HTTPException(status_code=400, detail="Missing filename")
    if len(raw) > MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="File exceeds 10MB limit")
    lower = file.filename.lower()
    buf = BytesIO(raw)
    if lower.endswith(".csv"):
        df = pd.read_csv(buf)
    elif lower.endswith((".xlsx", ".xls")):
        df = _read_excel_with_header_detection(buf)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Upload .xlsx, .xls, or .csv")
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    if len(df) > MAX_ROWS:
        df = df.head(MAX_ROWS)
    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no transaction rows")

    logger.info("Parsed %s rows from %s", len(df), file.filename)
    cols = _detect_columns(df)
    if not cols.get("description") or not cols.get("amount"):
        logger.info("Column auto-detect incomplete, calling Claude column mapper")
        cols = _claude_map_columns(df, db=db, company_id=company_id)
    if not cols.get("description"):
        raise HTTPException(status_code=400, detail=f"Could not detect description column. Columns: {list(df.columns)}")

    transactions = _normalize_transactions(df, cols)
    if not transactions:
        raise HTTPException(status_code=400, detail="No valid transactions found in file")

    try:
        result = _run_and_persist_compliance(
            transactions=transactions,
            period=period,
            company_trn=company_trn,
            entity_type=entity_type,
            company_id=company_id,
            company_name=company.name,
            db=db,
            actor=actor,
            source="manual",
            df=df,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("VAT compliance analyse failed")
        raise HTTPException(status_code=500, detail=f"Compliance analysis failed: {exc}") from exc

    logger.info(
        "VAT compliance analyse complete review_id=%s issues=%s",
        result["review_id"],
        result.get("analysis", {}).get("executive_summary", {}).get("issues_count"),
    )
    return result


@router.post("/analyse-from-db")
async def analyse_compliance_from_db(
    period: str = Form(...),
    company_trn: str = Form(""),
    entity_type: str = Form("Mainland UAE"),
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    """Run compliance review on all classified transactions already in VAT Classifier."""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    transactions = transactions_for_compliance_from_db(db, company_id)
    if not transactions:
        raise HTTPException(
            status_code=400,
            detail="No classified transactions found. Upload via Smart Upload or VAT Classifier first.",
        )
    try:
        return _run_and_persist_compliance(
            transactions=transactions,
            period=period,
            company_trn=company_trn or (company.trn or ""),
            entity_type=entity_type,
            company_id=company_id,
            company_name=company.name,
            db=db,
            actor="user",
            source="smart_upload",
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("VAT compliance analyse-from-db failed")
        raise HTTPException(status_code=500, detail=f"Compliance analysis failed: {exc}") from exc


@router.post("/draft-disclosure")
async def draft_disclosure(
    body: DraftDisclosureRequest,
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
):
    if claude_client is None:
        raise HTTPException(status_code=503, detail="Claude API not configured")
    prompt = f"""Draft a formal FTA voluntary disclosure letter for:
Company: {body.company_name}
TRN: {body.company_trn}
Period: {body.period}
Findings: {body.findings_summary}

Return ONLY JSON: {{ "letter_text": "full letter body as plain text" }}
Use professional formal tone suitable for FTA submission. Reference specific findings and corrected VAT amounts."""
    try:
        msg = claude_client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        data = _json_from_claude_text(msg.content[0].text)
        letter_text = str(data.get("letter_text", ""))
        try:
            log_ai_audit(
                db,
                company_id=company_id,
                user_email="user",
                action_type="ai_call",
                feature="vat_compliance_review",
                input_summary=f"Draft disclosure for {body.company_name} {body.period}",
                output_summary=letter_text[:100],
                status="success",
            )
        except Exception:
            pass
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Disclosure letter generation failed: {exc}") from exc

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(pdf_buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    flow: List[Any] = []
    for para in letter_text.split("\n\n"):
        if para.strip():
            flow.append(Paragraph(_pdf_text(para).replace("\n", "<br/>"), styles["Normal"]))
            flow.append(Spacer(1, 3 * mm))
    doc.build(flow)
    return {
        "letter_text": letter_text,
        "letter_pdf_base64": base64.b64encode(pdf_buf.getvalue()).decode("ascii"),
        "review_id": body.review_id,
    }


@router.get("/latest")
async def latest_compliance_review(
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
    include_exports: bool = Query(default=False),
):
    """Most recent compliance review for this company (e.g. from Smart Upload)."""
    row = (
        db.query(VATComplianceReview)
        .filter(VATComplianceReview.company_id == company_id)
        .order_by(VATComplianceReview.created_at.desc())
        .first()
    )
    if not row:
        return {"review": None}
    company = db.query(Company).filter(Company.id == company_id).first()
    analysis = _normalize_analysis(row.analysis or {})
    pdf_b64 = excel_b64 = ""
    if include_exports:
        pdf_b64, excel_b64 = _export_compliance_reports(
            analysis,
            company.name if company else "Company",
            row.company_trn or "",
            row.period or "",
            row.entity_type or "Mainland UAE",
            [],
            pd.DataFrame(),
        )
    return {
        "review": {
            "review_id": row.id,
            "period": row.period,
            "compliance_rating": row.compliance_rating,
            "issues_count": row.issues_count,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "source": row.source,
        },
        "analysis": analysis,
        "row_count": row.row_count,
        "pdf_base64": pdf_b64,
        "excel_base64": excel_b64,
        "review_id": row.id,
    }


@router.get("/history")
async def compliance_history(
    company_id: int = Depends(get_current_company_id),
    db: Session = Depends(get_db),
    limit: int = Query(default=10, ge=1, le=50),
):
    local_rows = (
        db.query(VATComplianceReview)
        .filter(VATComplianceReview.company_id == company_id)
        .order_by(VATComplianceReview.created_at.desc())
        .limit(limit)
        .all()
    )
    reviews = [
        {
            "id": r.id,
            "period": r.period,
            "compliance_rating": r.compliance_rating,
            "issues_count": r.issues_count,
            "net_vat_position": r.net_vat_position,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "source": r.source,
        }
        for r in local_rows
    ]
    if reviews:
        return {"reviews": reviews}
    sb = _get_supabase()
    if not sb:
        return {"reviews": []}
    try:
        res = (
            sb.table("vat_compliance_reviews")
            .select("id, period, compliance_rating, issues_count, net_vat_position, created_at")
            .eq("company_id", company_id)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
        )
        return {"reviews": res.data or []}
    except Exception as exc:
        logger.warning("Compliance history fetch failed: %s", exc)
        return {"reviews": []}
